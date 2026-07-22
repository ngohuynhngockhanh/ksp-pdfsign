"""Doc + cham loi file BCT (To khai 01/GTGT + 2 bang ke) ke toan up len.

Khong hardcode toa do o — nhan dien chi tieu theo nhan [NN] va cot thue suat
theo mode, de ben voi truong hop file bi xe dich dong/cot.
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime

from openpyxl import load_workbook

from . import tax_policy

# Chi tieu chi hop le khi O chi chua dung nhan [NN] (khong phai o cong thuc
# "[27]=[29]+[30]..." — o do co nhieu nhan long nhau, KHONG phai gia tri that).
_CT_FULL = re.compile(r"^\[(\d+[a-z]?)\]$")
_RATES = {0.0, 0.05, 0.08, 0.1}  # thue suat hop le (dang phan so)


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _fmt(v) -> str:
    """Gia tri o -> chuoi hien thi (so co dau phan cach, ngay dd/mm/yyyy)."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, bool):
        return "x" if v else ""
    if isinstance(v, (int, float)):
        r = round(v, 2)
        if abs(r - round(r)) < 1e-9:
            return f"{int(round(r)):,}".replace(",", ".")
        s = f"{r:,.2f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")
        return s
    return str(v)


# ---------------------------------------------------------------- To khai

def parse_to_khai(ws) -> dict[str, float | None]:
    """Tra ve {ma_chi_tieu: gia_tri}. Gia tri = so dau tien ben PHAI nhan,
    truoc nhan ke tiep cung dong."""
    ct: dict[str, float | None] = {}
    for row in ws.iter_rows():
        labels = []  # (cot, ma)
        for c in row:
            if isinstance(c.value, str):
                m = _CT_FULL.match(c.value.strip())
                if m:
                    labels.append((c.column, m.group(1)))
        if not labels:
            continue
        labels.sort()
        cells = sorted(((c.column, c.value) for c in row), key=lambda x: x[0])
        for i, (col, ma) in enumerate(labels):
            nxt = labels[i + 1][0] if i + 1 < len(labels) else 10**9
            val = None
            for cc, cv in cells:
                if col < cc < nxt and _num(cv) is not None:
                    val = float(cv)
                    break
            ct[ma] = val
    return ct


# ------------------------------------------------------------- Bang ke

def parse_bang_ke(ws) -> list[dict]:
    """Moi dong HD: {ngay, so_hd, ten, doanh_thu, rate, tien_thue}.

    Nhan dien cot thue suat = cot xuat hien gia tri {0.05,0.08,0.1} nhieu nhat;
    dong du lieu = dong co so o cot do VA co so tien ben trai."""
    from collections import Counter

    col_hits: Counter = Counter()
    for row in ws.iter_rows():
        for c in row:
            n = _num(c.value)
            if n is not None and n in {0.05, 0.08, 0.1}:
                col_hits[c.column] += 1
    if not col_hits:
        return []
    rate_col = col_hits.most_common(1)[0][0]

    rows: list[dict] = []
    for row in ws.iter_rows():
        by_col = {c.column: c.value for c in row}
        rv = by_col.get(rate_col)
        rate = _num(rv)
        if rate is None:
            continue
        # doanh thu = so gan nhat ben TRAI cot thue suat
        left = [(cc, _num(cv)) for cc, cv in by_col.items() if cc < rate_col and _num(cv) is not None]
        if not left:
            continue  # dong tieu de / tong -> bo
        dt_col, doanh_thu = max(left, key=lambda x: x[0])
        if doanh_thu is None or doanh_thu <= 0:
            continue
        right = [(cc, _num(cv)) for cc, cv in by_col.items() if cc > rate_col and _num(cv) is not None]
        tien_thue = min(right, key=lambda x: x[0])[1] if right else 0.0
        # ten doi tac = o CHU dai nhat ben trai cot doanh thu (khong phai ngay)
        texts = [cv for cc, cv in by_col.items() if cc < dt_col and isinstance(cv, str) and not _CT_FULL.match(cv.strip())]
        ten = max((t for t in texts), key=len, default="")
        ngay = next((cv.strftime("%Y-%m-%d") for cv in by_col.values() if isinstance(cv, datetime)), "")
        date_col = next((cc for cc, cv in by_col.items() if isinstance(cv, datetime)), None)
        if not ngay:
            date_hit = next(((cc, cv) for cc, cv in by_col.items() if isinstance(cv, str) and re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", cv.strip())), None)
            raw_date = date_hit[1] if date_hit else ""
            date_col = date_hit[0] if date_hit else None
            if raw_date:
                try:
                    ngay = datetime.strptime(raw_date.strip(), "%d/%m/%Y").strftime("%Y-%m-%d")
                except ValueError:
                    pass
        so_hd = ""
        if date_col:
            candidates = [(cc, cv) for cc, cv in by_col.items() if cc < date_col and isinstance(cv, (int, float, str))]
            if candidates:
                raw_so = max(candidates, key=lambda x: x[0])[1]
                so_hd = str(int(raw_so)) if isinstance(raw_so, float) and raw_so.is_integer() else str(raw_so).strip()
        rows.append({
            "doanh_thu": doanh_thu,
            "rate": rate,
            "rate_pct": round(rate * 100),
            "tien_thue": tien_thue or 0.0,
            "ten": ten.strip(),
            "ngay": ngay,
            "so_hd": so_hd,
        })
    return rows


# --------------------------------------------------------------- Grid

def sheet_grid(ws) -> dict:
    rows = []
    for row in ws.iter_rows():
        rows.append([_fmt(c.value) for c in row])
    # bo cac dong trong o cuoi
    while rows and not any(x.strip() for x in rows[-1]):
        rows.pop()
    ncols = max((len(r) for r in rows), default=0)
    for r in rows:
        r += [""] * (ncols - len(r))
    merges = []
    for mr in ws.merged_cells.ranges:
        merges.append([mr.min_row - 1, mr.min_col - 1, mr.max_row - 1, mr.max_col - 1])
    return {"name": ws.title, "rows": rows, "ncols": ncols, "merges": merges}


# --------------------------------------------------------------- Cham loi

def _g(ct, *keys):
    for k in keys:
        v = ct.get(k)
        if v is not None:
            return v
    return 0.0


def check(ct: dict, ban: list[dict], mua: list[dict]) -> list[dict]:
    """Tra ve list finding: {level: 'do'|'vang', title, detail, cells:[ma...]}."""
    out: list[dict] = []

    def add(level, title, detail, cells=None):
        out.append({"level": level, "title": title, "detail": detail, "cells": cells or []})

    c22 = _g(ct, "22"); c25 = _g(ct, "25"); c35 = _g(ct, "35", "28")
    c36 = _g(ct, "36"); c37 = _g(ct, "37"); c38 = _g(ct, "38"); c39 = _g(ct, "39")
    c40a = ct.get("40a"); c40 = ct.get("40")

    # 1) [40a]/[40] khong duoc am
    tmp = c36 - c22 + c37 - c38 + c39
    am = [ma for ma, val in (("40a", c40a), ("40", c40)) if val is not None and val < -0.5]
    if am:
        carry = max(0.0, -tmp)
        add("do", "Kỳ này chưa phải nộp VAT — số âm đang đặt sai chỉ tiêu",
            f"Kết quả tính là {_fmt(tmp)} < 0, nghĩa là doanh nghiệp còn dư {_fmt(carry)} thuế GTGT "
            f"đầu vào được khấu trừ và kỳ này chưa phải nộp VAT. Tuy nhiên tờ khai không để số âm ở "
            f"[40a]/[40]: cần ghi [40a]=[40]=0 và chuyển {_fmt(carry)} sang [41]/[43] để chuyển kỳ sau.",
            am + ["41", "43"])

    # 2) Kiem tra cong thuc noi bo
    def formula(ma, expect, comps):
        got = ct.get(ma)
        if got is not None and abs(got - expect) > 1:
            add("do", f"[{ma}] lệch công thức", f"[{ma}]={_fmt(got)} nhưng {comps}={_fmt(expect)}.", [ma])
    formula("36", c35 - c25, "[35]−[25]")
    c26 = _g(ct, "26"); c27 = _g(ct, "27")
    formula("28", _g(ct, "31") + _g(ct, "33"), "[31]+[33]")
    formula("34", c26 + c27, "[26]+[27]")

    # 3) Doi chieu bang ke <-> to khai
    sum_ban_dt = sum(r["doanh_thu"] for r in ban)
    sum_ban_thue = sum(r["tien_thue"] for r in ban)
    sum_mua_dt = sum(r["doanh_thu"] for r in mua)
    sum_mua_thue = sum(r["tien_thue"] for r in mua)
    tk_ban_dt = _g(ct, "34", "27"); tk_ban_thue = _g(ct, "35", "28")
    tk_mua_dt = _g(ct, "23"); tk_mua_thue = _g(ct, "24", "25")
    if ban and abs(sum_ban_dt - tk_ban_dt) > 1:
        add("do", "Doanh thu bảng kê bán ra ≠ tờ khai",
            f"Tổng bảng kê = {_fmt(sum_ban_dt)}, tờ khai [34]/[27] = {_fmt(tk_ban_dt)} (lệch {_fmt(sum_ban_dt - tk_ban_dt)}).", ["34", "27"])
    if ban and abs(sum_ban_thue - tk_ban_thue) > 1:
        add("do", "Thuế bảng kê bán ra ≠ tờ khai",
            f"Tổng thuế bảng kê = {_fmt(sum_ban_thue)}, tờ khai [35]/[28] = {_fmt(tk_ban_thue)}.", ["35", "28"])
    if mua and abs(sum_mua_dt - tk_mua_dt) > 1:
        add("do", "Giá trị mua vào bảng kê ≠ tờ khai",
            f"Tổng bảng kê = {_fmt(sum_mua_dt)}, tờ khai [23] = {_fmt(tk_mua_dt)}.", ["23"])
    if mua and abs(sum_mua_thue - tk_mua_thue) > 1:
        add("do", "Thuế mua vào bảng kê ≠ tờ khai",
            f"Tổng thuế bảng kê = {_fmt(sum_mua_thue)}, tờ khai [24]/[25] = {_fmt(tk_mua_thue)}.", ["24", "25"])

    # 4) HD ban thue suat 0% cho DN noi dia
    zero = [r for r in ban if r["rate_pct"] == 0 and r["doanh_thu"] > 0]
    dn_noi_dia = [r for r in zero if "CÔNG TY" in (r["ten"] or "").upper() or "CTY" in (r["ten"] or "").upper()]
    if dn_noi_dia:
        tong = sum(r["doanh_thu"] for r in dn_noi_dia)
        c29 = ct.get("29")
        detail = (f"{len(dn_noi_dia)} HĐ đang ghi 0%, tổng doanh thu {_fmt(tong)}. Cần phân loại theo bản chất: "
                  "nếu là sản phẩm/dịch vụ phần mềm thuộc diện không chịu thuế thì khai KCT tại [26], không phải 0%; "
                  "nếu là xuất khẩu đủ điều kiện 0% thì khai [29]; nếu là hàng hóa/dịch vụ trong nước vốn chịu 10% "
                  "và đủ điều kiện giảm trong giai đoạn 01/07/2025–31/12/2026 thì áp 8%. ")
        if not c29 or c29 < 1:
            detail += " Hiện [29] đang trống."
        add("vang", "Dòng 0% cần xác nhận: phần mềm KCT, xuất khẩu 0%, hay hàng chịu 8%", detail,
            ["29"] + [f"__ten:{r['ten']}" for r in dn_noi_dia[:8]])

    # 4b) Nhom [32]/[33] ghi 10% nhung thue thuc te khac (lan HD 0%/8%)
    c32 = _g(ct, "32"); c33 = _g(ct, "33")
    if c32 > 0 and c33 > 0:
        implied = c33 / c32 * 100
        base_8 = sum(r["doanh_thu"] for r in ban if r["rate_pct"] == 8)
        base_0 = sum(r["doanh_thu"] for r in ban if r["rate_pct"] == 0)
        if abs(implied - 8) > 0.25 and base_8 and abs(c32 - base_8 - base_0) <= 2:
            add("vang", f"Tỷ lệ {implied:.2f}% không phải thuế suất — [32] đang gộp sai nhóm",
                f"[32] gồm {_fmt(base_8)} doanh thu 8% + {_fmt(base_0)} doanh thu đang ghi 0%, nhưng [33] "
                f"chỉ có {_fmt(c33)} thuế của phần 8%; vì vậy phép chia ra {implied:.2f}%. "
                f"Sau khi tách {_fmt(base_0)} sang [26] nếu là phần mềm KCT, hoặc [29] nếu thật sự đủ điều kiện 0%, "
                f"[32] còn {_fmt(base_8)} và [33]/[32] sẽ xấp xỉ 8%.", ["32", "33", "26", "29"])

    # 5) Thue suat la (suy ra) khong thuoc {0,5,8,10}
    la = sorted({r["rate_pct"] for r in (ban + mua) if r["rate_pct"] not in (0, 5, 8, 10)})
    if la:
        add("vang", "Có thuế suất lạ", f"Xuất hiện thuế suất {', '.join(str(x) + '%' for x in la)} — kiểm tra lại hóa đơn.", [])

    # 6) Checkpoint thue theo ngay hoa don, khong ket luan 10% sai vi co nhom loai tru.
    outside_8 = []
    verify_10 = []
    for r in ban:
        try:
            on_date = date.fromisoformat(r.get("ngay") or "")
        except ValueError:
            continue
        active = tax_policy.reduction_active(on_date)
        if r["rate_pct"] == 8 and not active:
            outside_8.append(r)
        elif r["rate_pct"] == 10 and active:
            verify_10.append(r)
    if outside_8:
        add("do", "Có hóa đơn 8% nằm ngoài các giai đoạn giảm thuế đã cấu hình",
            f"Phát hiện {len(outside_8)} dòng 8% ngoài các checkpoint giảm thuế từ năm 2022 đến hết 2026; cần kiểm tra chính sách có hiệu lực tại đúng ngày hóa đơn.", [])
    if verify_10:
        add("vang", "Có dòng 10% trong thời gian đang giảm 2% — kiểm tra nhóm loại trừ",
            f"Có {len(verify_10)} dòng 10% trong giai đoạn 01/07/2025–31/12/2026. 10% vẫn đúng nếu hàng hóa/dịch vụ thuộc nhóm loại trừ; nếu không, mức áp dụng là 8%.", [])

    return out


def crosscheck_sales(db, result: dict) -> dict:
    """Dung XML hoa don trong CRM de xac nhan cac dong 0% thuc chat la KCT."""
    from sqlalchemy import select

    from .db import InvSale

    confirmed = []
    for row in result.get("ban", []):
        if row.get("rate_pct") != 0 or not row.get("so_hd"):
            continue
        candidates = list(db.scalars(select(InvSale).where(InvSale.so_hd == row["so_hd"])))
        inv = next((x for x in candidates if not row.get("ngay") or x.ngay == row["ngay"]), None)
        if inv and inv.lines and all(line.thue_kct for line in inv.lines):
            confirmed.append((row, inv))
    if not confirmed:
        return result
    total = sum(row["doanh_thu"] for row, _ in confirmed)
    numbers = ", ".join(str(inv.so_hd) for _, inv in confirmed)
    for finding in result["findings"]:
        if finding["title"].startswith("Dòng 0% cần xác nhận"):
            finding["title"] = "Đã xác nhận là phần mềm KCT — bảng kê đang xếp sai nhóm"
            finding["detail"] = (
                f"XML trong CRM xác nhận HĐ {numbers} đều là sản phẩm/dịch vụ phần mềm và các dòng đã đánh dấu "
                f"không chịu thuế (KCT), tổng {_fmt(total)}. Cần chuyển doanh thu này khỏi nhóm [32] sang [26]; "
                "không khai ở [29] và không tính thuế 8%."
            )
            finding["cells"] = ["26", "29", "32"]
        elif "không phải thuế suất" in finding["title"]:
            finding["detail"] += f" CRM đã xác nhận {_fmt(total)} này là phần mềm KCT, nên phương án đúng là chuyển sang [26]."
    return result


# --------------------------------------------------------------- Entry

def review_bytes(xlsx_bytes: bytes) -> dict:
    """Parse + cham loi 1 file BCT. Tra {ct, ban, mua, findings, grids, summary}."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    def find_sheet(*kw):
        for ws in wb.worksheets:
            t = ws.title.lower()
            if all(k in t for k in kw):
                return ws
        return None

    ws_tk = find_sheet("khai") or wb.worksheets[0]
    ws_ban = find_sheet("bán") or find_sheet("ban")
    ws_mua = find_sheet("mua")

    ct = parse_to_khai(ws_tk)
    ban = parse_bang_ke(ws_ban) if ws_ban else []
    mua = parse_bang_ke(ws_mua) if ws_mua else []
    findings = check(ct, ban, mua)
    grids = [sheet_grid(ws) for ws in wb.worksheets]

    summary = {
        "ban_ra_dt": _g(ct, "34", "27"), "ban_ra_thue": _g(ct, "35", "28"),
        "mua_vao_dt": _g(ct, "23"), "mua_vao_thue": _g(ct, "24", "25"),
        "khau_tru_ky_truoc": _g(ct, "22"),
        "ct_36": ct.get("36"),
        "ct_40": max(0.0, ct.get("40", ct.get("40a")) or 0.0),
        "ct_41": max(ct.get("41") or 0.0, max(0.0, -(_g(ct, "36") - _g(ct, "22") + _g(ct, "37") - _g(ct, "38") + _g(ct, "39")))),
        "ct_43": max(ct.get("43") or 0.0, max(0.0, -(_g(ct, "36") - _g(ct, "22") + _g(ct, "37") - _g(ct, "38") + _g(ct, "39")))),
        "so_hd_ban": len(ban), "so_hd_mua": len(mua),
        "do": sum(1 for f in findings if f["level"] == "do"),
        "vang": sum(1 for f in findings if f["level"] == "vang"),
    }
    return {"ct": ct, "ban": ban, "mua": mua, "findings": findings, "grids": grids, "summary": summary}
