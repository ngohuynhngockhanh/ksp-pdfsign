"""Xuat Excel/ZIP cho danh sach ton kho (mua vao / ban ra / xuat kho / san xuat)."""
from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

_IHOADON_TEMPLATE = Path(__file__).resolve().parent / "assets" / "ihoadon_template.xlsx"
# Cot (1-indexed) trong sheet "Bang_ke_hang_hoa_dich_vu" theo field ky thuat hang 10
_IH_COL = {
    "view_order": 1,       # STT
    "product_code": 3,     # ma hang
    "product_name": 4,     # ten hang
    "unit_name": 10,       # DVT
    "quantity": 13,        # so luong
    "price": 14,           # don gia
    "amount_wo_disc": 15,  # thanh tien chua tru CK
    "amount": 18,          # thanh tien
    "vat_name": 19,        # thue suat (chuoi, vd '8%','10%x70%','Không chịu thuế')
    "amount_vat": 20,      # tien thue
    "is_money_service": 25,  # 'x' neu la phi dich vu
}
_IH_DATA_ROW0 = 11  # dong dau tien dien du lieu


def export_ihoadon_xlsx(lines: list[dict]) -> io.BytesIO:
    """Dien danh sach dong hang vao khuon 'Bảng kê hàng hóa dịch vụ' cua iHoadon.

    Giu nguyen header + sheet Danh_muc (dropdown validate) tu template. So HD/ky
    hieu de TRONG (dien tay tren iHoadon). Moi dong dict:
      {ma_hang, ten, dvt, so_luong, don_gia, thanh_tien, vat_name, tien_thue, is_dich_vu}
    """
    wb = load_workbook(_IHOADON_TEMPLATE)
    ws = wb["Bang_ke_hang_hoa_dich_vu"]
    r = _IH_DATA_ROW0
    for i, ln in enumerate(lines, start=1):
        ws.cell(row=r, column=_IH_COL["view_order"], value=i)
        ws.cell(row=r, column=_IH_COL["product_code"], value=ln.get("ma_hang") or "")
        ws.cell(row=r, column=_IH_COL["product_name"], value=ln.get("ten") or "")
        ws.cell(row=r, column=_IH_COL["unit_name"], value=ln.get("dvt") or "")
        sl = float(ln.get("so_luong") or 0)
        dg = float(ln.get("don_gia") or 0)
        tt = float(ln.get("thanh_tien") or 0) or round(sl * dg)
        ws.cell(row=r, column=_IH_COL["quantity"], value=sl)
        ws.cell(row=r, column=_IH_COL["price"], value=dg)
        ws.cell(row=r, column=_IH_COL["amount_wo_disc"], value=tt)
        ws.cell(row=r, column=_IH_COL["amount"], value=tt)
        ws.cell(row=r, column=_IH_COL["vat_name"], value=ln.get("vat_name") or "")
        if ln.get("tien_thue") is not None:
            ws.cell(row=r, column=_IH_COL["amount_vat"], value=float(ln.get("tien_thue") or 0))
        if ln.get("is_dich_vu"):
            ws.cell(row=r, column=_IH_COL["is_money_service"], value="x")
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def ihoadon_response(lines: list[dict], filename: str) -> StreamingResponse:
    buf = export_ihoadon_xlsx(lines)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{_ascii_filename(filename)}"'},
    )

_BAD_FS_CHARS = re.compile(r'[\\/:*?"<>|]')


def sanitize_arcname(name: str) -> str:
    """Bo ky tu cam trong ten file he thong (giu dau tieng Viet)."""
    return _BAD_FS_CHARS.sub("_", name).strip() or "file"


def _ascii_filename(name: str) -> str:
    """Content-Disposition an toan (bo dau, giu duoi file) - giong _content_disposition o main.py."""
    return name.encode("ascii", "ignore").decode() or "export"


def xlsx_response(sheets: list[tuple[str, list[str], list[list]]], filename: str) -> StreamingResponse:
    """sheets: [(ten_sheet, headers, rows), ...] -> file .xlsx (nhieu sheet)."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, headers, rows in sheets:
        ws = wb.create_sheet(title=name[:31])  # excel gioi han 31 ky tu/sheet
        ws.append(headers)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{_ascii_filename(filename)}"'},
    )


def zip_response(files: list[tuple[str, bytes]], filename: str) -> StreamingResponse:
    """files: [(arcname, content), ...] -> zip trong bo nho. Ten trung -> them hau to ' (2)', ' (3)'..."""
    buf = io.BytesIO()
    seen: dict[str, int] = {}
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, content in files:
            if name in seen:
                seen[name] += 1
                stem, dot, ext = name.rpartition(".")
                arc = f"{stem} ({seen[name]}){dot}{ext}" if dot else f"{name} ({seen[name]})"
            else:
                seen[name] = 1
                arc = name
            zf.writestr(arc, content)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{_ascii_filename(filename)}"'},
    )
