"""Import du lieu ton kho: file Excel ton dau ky + hoa don mua vao.

Excel "TONG HOP TON KHO" (xuat tu phan mem ke toan): cac section
"Tên kho : ..." roi den dong hang [_, Ma hang, Ten, DVT, dau_ky SL/GT,
nhap SL/GT, xuat SL/GT, cuoi_ky SL/GT]. Chi lay CUOI KY -> ton dau ky nam sau.
Mot ma hang co the nam o nhieu kho -> 1 InvItem, ton tach theo kho.
"""
from __future__ import annotations

import difflib
import io
import json
import re
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote, urlparse
import xml.etree.ElementTree as ET

from sqlalchemy import select
from sqlalchemy.orm import Session

import httpx

from . import accounts, ai, invoice, money
from .config import Settings
from .db import (
    InvItem,
    InvItemAlias,
    InvMove,
    InvPurchase,
    InvPurchaseLine,
    InvSale,
    InvSaleLine,
    InvWarehouse,
)
from .inventory import normalize_name, normalize_so_hd

# Nhan dien kho trong Excel theo ten (bo dau, chu thuong)
_WAREHOUSE_KEYWORDS = [
    ("nguyen vat lieu", "NVL"),
    ("thanh pham", "TP"),
    ("hang hoa", "HH"),
]

OPENING_REF = "opening"

# Tu khoa doan hoa don DICH VU/chi phi (khong nhap kho). Da chuan hoa (bo dau,
# thuong). Chi cac truong hop RO RANG — tranh gan nham NCC ban hang hoa.
# Vi du KHONG dung 'dich vu' vi nhieu NCC hang hoa co ten '... DICH VU DIEN TU'.
_DICH_VU_KEYWORDS = (
    "be group", "grab", "gojek", "go jek", "baemin",
    "dien luc", "vien thong", "viettel", "vnpt", "mobifone", "fpt telecom",
    "quang cao", "google", "facebook", "meta platforms",
    "van tai", "van chuyen", "giao hang", "chuyen phat", "buu chinh", "logistics",
    "bao hiem", "internet", "cuoc xe", "cuoc taxi", "tien dien", "tien nuoc",
    "am thuc", "nha hang", "quan an", "cafe", "ca phe", "khach san", "an uong",
    "nha nghi", "du lich", "lu hanh", "ve may bay", "taxi",
    "van chuyen", "cuoc van chuyen", "phi van chuyen", "phi ship", "phi giao hang",
    "boc xep", "kho bai", "phi dich vu", "phi tu van",
)

# Dong co dinh dang "so + don vi dong goi" (vd '500 gr', '1 thung', '2 cai') ->
# gan nhu chac chan la HANG HOA, KHONG phai dich vu — dung de chan doan nham
# dich_vu khi ten NCC/dong hang trung tu khoa dich vu mot cach tinh co.
_HANG_HOA_UNIT_RE = re.compile(
    r"\b\d+\s*(gr|g|kg|ml|lit|l|goi|hop|thung|chai|lon|cai|bo)\b"
)


def _wh_code_from_name(name: str) -> str | None:
    n = normalize_name(name)
    for kw, code in _WAREHOUSE_KEYWORDS:
        if kw in n:
            return code
    return None


def _f(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_opening_xlsx(data: bytes) -> dict:
    """Doc file Excel ton kho -> danh sach mat hang + ton cuoi ky + canh bao.

    Tra ve:
      items: {ma_hang: {ten, dvt, stocks: {wh_code: {sl, gt}}}}
      warnings: [{code, msg}]
      tong: {so_ma, so_ma_ton, tong_sl, tong_gia_tri}
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.worksheets[0]

    items: dict[str, dict] = {}
    warnings: list[dict] = []
    cur_wh: str | None = None
    n_rows = 0

    for row in ws.iter_rows(values_only=True):
        first = str(row[0]).strip() if row[0] is not None else ""
        if first.lower().startswith("tên kho") or first.lower().startswith("ten kho"):
            code = _wh_code_from_name(first)
            if code is None:
                warnings.append({
                    "code": "kho_la",
                    "msg": f"Không nhận diện được kho '{first}' — bỏ qua các dòng thuộc kho này",
                })
            cur_wh = code
            continue
        ma = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not ma or ma.lower().startswith("số dòng") or ma.lower().startswith("so dong"):
            continue
        if cur_wh is None:
            continue
        n_rows += 1
        ten = str(row[2]).strip() if row[2] is not None else ""
        dvt = str(row[3]).strip() if row[3] is not None else ""
        sl_cuoi = _f(row[10]) if len(row) > 10 else 0.0
        gt_cuoi = _f(row[11]) if len(row) > 11 else 0.0

        it = items.get(ma)
        if it is None:
            it = {"ma_hang": ma, "ten": ten, "dvt": dvt, "stocks": {}}
            items[ma] = it
        else:
            # Cung ma o 2 kho: doi chieu ten/DVT, uu tien gia tri khong rong
            if ten and it["ten"] and normalize_name(ten) != normalize_name(it["ten"]):
                warnings.append({
                    "code": "ten_lech",
                    "msg": f"Mã {ma}: tên khác nhau giữa 2 kho ('{it['ten']}' ≠ '{ten}') — lấy tên đầu tiên",
                })
            if not it["ten"]:
                it["ten"] = ten
            if not it["dvt"]:
                it["dvt"] = dvt
        if cur_wh in it["stocks"]:
            warnings.append({
                "code": "trung_dong",
                "msg": f"Mã {ma}: xuất hiện 2 lần trong cùng kho {cur_wh} — cộng dồn",
            })
            it["stocks"][cur_wh]["sl"] += sl_cuoi
            it["stocks"][cur_wh]["gt"] += gt_cuoi
        else:
            it["stocks"][cur_wh] = {"sl": sl_cuoi, "gt": gt_cuoi}

    if not items:
        raise ValueError(
            "Không đọc được dòng hàng nào — file có đúng định dạng 'TỔNG HỢP TỒN KHO' không?"
        )

    # Canh bao chat luong du lieu
    tong_sl = 0.0
    tong_gt = 0.0
    so_ma_ton = 0
    for ma, it in items.items():
        if not it["dvt"]:
            warnings.append({"code": "thieu_dvt", "msg": f"Mã {ma} ({it['ten'][:40]}): thiếu ĐVT"})
        has_ton = False
        for wh, s in it["stocks"].items():
            sl, gt = s["sl"], s["gt"]
            if sl == 0 and gt != 0:
                warnings.append({
                    "code": "gia_tri_treo",
                    "msg": (
                        f"Mã {ma} ({it['ten'][:40]}) kho {wh}: SL=0 nhưng giá trị "
                        f"{gt:,.0f}đ — KHÔNG import, cần anh và kế toán xử lý "
                        f"(điều chỉnh giá trị treo)"
                    ),
                })
                s["skip"] = True
            elif sl < 0 or gt < 0:
                warnings.append({
                    "code": "ton_am",
                    "msg": f"Mã {ma} kho {wh}: tồn cuối kỳ âm (SL={sl}, GT={gt}) — KHÔNG import",
                })
                s["skip"] = True
            elif sl > 0:
                if gt == 0:
                    warnings.append({
                        "code": "gia_von_0",
                        "msg": f"Mã {ma} ({it['ten'][:40]}) kho {wh}: SL={sl:g} nhưng giá trị 0đ — import với giá vốn 0",
                    })
                has_ton = True
                tong_sl += sl
                tong_gt += gt
        if has_ton:
            so_ma_ton += 1

    return {
        "items": items,
        "warnings": warnings,
        "tong": {
            "so_ma": len(items),
            "so_dong": n_rows,
            "so_ma_ton": so_ma_ton,
            "tong_sl": tong_sl,
            "tong_gia_tri": tong_gt,
        },
    }


def apply_opening(db: Session, parsed: dict, ngay: str = "2025-12-31") -> dict:
    """Ghi ton dau ky vao so: upsert InvItem + tao move dau_ky.

    Chi cho phep khi so kho chua co phat sinh (ngoai dau_ky cu — se bi thay the).
    """
    has_other = db.scalar(
        select(InvMove.id).where(InvMove.loai != "dau_ky").limit(1)
    )
    if has_other is not None:
        raise PermissionError(
            "Sổ kho đã có phát sinh (nhập/xuất) — không thể import lại tồn đầu kỳ. "
            "Cần xóa/void các chứng từ trước."
        )

    wh_map = {w.code: w for w in db.scalars(select(InvWarehouse))}
    existing = {i.ma_hang: i for i in db.scalars(select(InvItem))}

    # Thay the dau_ky cu (import lai)
    old = list(db.scalars(select(InvMove).where(InvMove.loai == "dau_ky")))
    for m in old:
        db.delete(m)

    n_items_new = 0
    n_moves = 0
    for ma, it in parsed["items"].items():
        # Ghi chu canh bao vao mat hang de duyet tay trong app
        notes = []
        if not it["dvt"]:
            notes.append("thiếu ĐVT")
        for wh_code, s in it["stocks"].items():
            if s.get("skip") and s["gt"]:
                notes.append(
                    f"giá trị treo {s['gt']:,.0f}đ kho {wh_code} (SL=0) CHƯA import — hỏi kế toán xử lý điều chỉnh"
                )
            elif s["sl"] > 0 and s["gt"] == 0:
                notes.append(f"tồn {s['sl']:g} kho {wh_code} nhưng giá vốn 0đ")
        auto_note = ("⚠️ " + "; ".join(notes)) if notes else ""

        obj = existing.get(ma)
        if obj is None:
            obj = InvItem(
                ma_hang=ma,
                ten=it["ten"],
                ten_norm=normalize_name(it["ten"]),
                dvt=it["dvt"],
                note=auto_note,
            )
            db.add(obj)
            existing[ma] = obj
            n_items_new += 1
        else:
            obj.ten = it["ten"] or obj.ten
            obj.ten_norm = normalize_name(obj.ten)
            obj.dvt = it["dvt"] or obj.dvt
            if auto_note and (not obj.note or obj.note.startswith("⚠️")):
                obj.note = auto_note
        for wh_code, s in it["stocks"].items():
            if s.get("skip") or s["sl"] <= 0:
                continue
            wh = wh_map.get(wh_code)
            if wh is None:
                continue
            db.flush()  # co obj.id
            db.add(InvMove(
                item_id=obj.id,
                warehouse_id=wh.id,
                ngay=ngay,
                loai="dau_ky",
                so_luong=s["sl"],
                don_gia=(s["gt"] / s["sl"]) if s["sl"] else 0.0,
                gia_tri=s["gt"],
                ref_type=OPENING_REF,
            ))
            n_moves += 1

    db.commit()
    return {
        "items_new": n_items_new,
        "items_total": len(parsed["items"]),
        "moves": n_moves,
        "replaced_old": len(old),
    }


# ---------------------------------------------------------------------------
# Hoa don mua vao -> draft (XML chinh xac > PDF pdfplumber > AI scan)
# ---------------------------------------------------------------------------
def _ngay_iso(d: dict | None) -> str:
    if not d:
        return ""
    try:
        return f"{int(d['year']):04d}-{int(d['month']):02d}-{int(d['day']):02d}"
    except (KeyError, TypeError, ValueError):
        return ""


def parse_purchase_xml(xml_bytes: bytes) -> dict:
    """Parse hoa don XML (TT78) theo goc NGUOI MUA: lay ben BAN (NBan) + tong tien.

    Dung lai parse_invoice_xml cho items; them NBan, SHDon, TToan.
    """
    base = invoice.parse_invoice_xml(xml_bytes)
    root = ET.fromstring(xml_bytes)

    def txt(el, tag, default=""):
        if el is None:
            return default
        x = el.find(tag)
        return (x.text or "").strip() if x is not None and x.text else default

    nban = root.find(".//NBan")
    shd = root.find(".//SHDon")
    tt = root.find(".//TToan")
    return {
        "source": "xml",
        "so_hd": (shd.text or "").strip() if shd is not None else "",
        "ky_hieu": base.get("ky_hieu", ""),
        "ngay": _ngay_iso(base.get("ngay")),
        "ten_ban": txt(nban, "Ten"),
        "mst_ban": txt(nban, "MST"),
        "tong_truoc_thue": money.parse_num(txt(tt, "TgTCThue")),
        "tong_thue": money.parse_num(txt(tt, "TgTThue")),
        "tong_tien": money.parse_num(txt(tt, "TgTTTBSo")),
        "items": base.get("items", []),
        "confidence": 1.0,
        "warnings": [],
    }


# ---------------------------------------------------------------------------
# Trich xuat ngay / so HD / ky hieu tu raw_text PDF (nhieu mau hoa don khac nhau)
# ---------------------------------------------------------------------------
_NGAY_SLASH_NEAR_DATE = re.compile(
    r"Ng[àa]y\s*\(?[Dd]ate\)?\s*:?\s*(\d{1,2})\s*[/-]\s*(\d{1,2})\s*[/-]\s*(\d{4})"
)
_NGAY_THANG_NAM = re.compile(
    r"[Nn]g[àa]y\s*(?:\([^)]{0,12}\)\s*)?:?\s*(\d{1,2})\s*"
    r"th[áa]ng\s*(?:\([^)]{0,12}\)\s*)?(\d{1,2})\s*"
    r"n[ăa]m\s*(?:\([^)]{0,12}\)\s*)?(\d{4})"
)
_NGAY_SLASH_LINE = re.compile(r"(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{4})")

_SO_HD_PATTERNS = [
    re.compile(r"Số\s*/?\(?(?:Invoice\s*)?(?:No\.?)?\)?\s*:?\s*(\d{1,8})"),
    re.compile(r"Số hóa đơn\s*:?\s*(\d{1,8})", re.IGNORECASE),
    re.compile(r"(?:^|\s)(?:HĐ|Invoice)\s*(?:số|No)\.?\s*:?\s*(\d{1,8})", re.IGNORECASE),
]

_KY_HIEU_RE = re.compile(r"\b([1-6]?[CK]2[0-9][A-Z]{2,5})\b")


def _valid_date(y: int, m: int, d: int) -> bool:
    return 2020 <= y <= 2035 and 1 <= m <= 12 and 1 <= d <= 31


def _mk_iso(y: str | int, m: str | int, d: str | int) -> str:
    return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"


def _extract_ngay(raw: str, base_ngay: dict | None) -> str:
    """Thu lan luot: (a) base parse, (b) Ngày(Date) dd/mm/yyyy, (c) ngày D tháng M năm Y,
    (d) dd/mm/yyyy tren cung dong co chu 'Ngày'."""
    iso = _ngay_iso(base_ngay)
    if iso:
        y, mo, d = (int(x) for x in iso.split("-"))
        if _valid_date(y, mo, d):
            return iso

    m = _NGAY_SLASH_NEAR_DATE.search(raw)
    if m:
        d, mo, y = m.groups()
        if _valid_date(int(y), int(mo), int(d)):
            return _mk_iso(y, mo, d)

    m = _NGAY_THANG_NAM.search(raw)
    if m:
        d, mo, y = m.groups()
        if _valid_date(int(y), int(mo), int(d)):
            return _mk_iso(y, mo, d)

    for line in raw.splitlines():
        low = line.lower()
        if "ngày" not in low and "ngay" not in low:
            continue
        m = _NGAY_SLASH_LINE.search(line)
        if m:
            d, mo, y = m.groups()
            if _valid_date(int(y), int(mo), int(d)):
                return _mk_iso(y, mo, d)
    return ""


def _extract_so_hd(raw: str) -> str:
    """Uu tien theo thu tu pattern; trong cac ket qua cua 1 pattern, chon so co
    nhieu chu so nhat (>=3 chu so), neu khong co thi lay ket qua dau tien."""
    for pat in _SO_HD_PATTERNS:
        cands = [mm.group(1) for mm in pat.finditer(raw)]
        if not cands:
            continue
        good = [c for c in cands if len(c) >= 3]
        return max(good, key=len) if good else cands[0]
    return ""


def _extract_ky_hieu(raw: str, base_ky_hieu: str) -> str:
    if base_ky_hieu:
        return base_ky_hieu
    m = _KY_HIEU_RE.search(raw)
    return m.group(1) if m else ""


def _pdf_sign_date(pdf_bytes: bytes) -> str:
    """Ngay ky so tren PDF (YYYY-MM-DD) neu co chu ky; '' neu khong."""
    try:
        import io as _io

        from pyhanko.pdf_utils.reader import PdfFileReader

        reader = PdfFileReader(_io.BytesIO(pdf_bytes))
        for sig in reader.embedded_signatures:
            dt = getattr(sig, "signer_reported_dt", None)
            if dt:
                return dt.strftime("%Y-%m-%d")
    except Exception:  # noqa: BLE001
        pass
    return ""


# Dong phan mem/license -> KCT (khong chiu thue GTGT)
_PHAN_MEM_RE = re.compile(r"phần mềm|phan mem|software|license|bản quyền|ban quyen", re.IGNORECASE)
_VAT_LEVELS = (0, 5, 8, 10)  # cac muc thue GTGT hop le


def _snap_vat(pct: float) -> int:
    """Lam tron ve muc thue GTGT gan nhat (0/5/8/10) — tong OCR hay lech nhe."""
    return min(_VAT_LEVELS, key=lambda v: abs(v - pct))


def _invoice_vat_rate(raw: str, tong_truoc_thue: float, tong_thue: float) -> int:
    """Thue suat muc hoa don: regex text > suy tu tong > mac dinh 8%."""
    m = re.search(r"[Tt]huế suất(?:\s*GTGT)?[^\d%]{0,15}(\d{1,2})\s*%", raw)
    if m:
        return int(m.group(1))
    if re.search(r"[Tt]huế suất[^\n]{0,20}(KCT|[Kk]hông chịu thuế)", raw):
        return 0
    if tong_truoc_thue > 0:
        return _snap_vat(tong_thue / tong_truoc_thue * 100)
    return 8  # khong suy duoc gi -> mac dinh 8% (quy uoc cong ty)


def _line_vat(ten: str, invoice_rate: int) -> int:
    """VAT cho 1 dong: phan mem/license -> KCT (0), con lai theo muc hoa don."""
    if _PHAN_MEM_RE.search(ten):
        return 0
    return invoice_rate


def parse_purchase_pdf(pdf_bytes: bytes) -> dict:
    """Parse hoa don PDF (co lop text): items tu bang + ben ban tu text tho."""
    base = invoice.parse_invoice(pdf_bytes)
    raw = base.get("raw_text", "")
    ngay_ky = _pdf_sign_date(pdf_bytes)

    # Ben ban: MST dau tien trong van ban (truoc block nguoi mua)
    ten_ban = ""
    mst_ban = ""
    m = re.search(r"Mã số thuế[^\d]*(\d{10,13})", raw)
    if m:
        mst_ban = m.group(1)
    # Ten don vi ban: uu tien dong chua tu dinh danh phap nhan (CONG TY/DN/HKD...).
    # Nguoi ban la ben KHONG phai INUT (MST mua = 4401053694).
    _ORG_KW = ("công ty", "cong ty", "cty", "doanh nghiệp", "hộ kinh doanh",
               "chi nhánh", "trung tâm", "hợp tác xã", "htx", "cửa hàng")
    _SKIP = ("hóa đơn", "mẫu số", "ký hiệu", "bản thể hiện", "invoice", "số:",
             "(buyer)", "người mua", "đơn vị mua")
    cand = ""
    for line in raw.splitlines():
        t = line.strip()
        low = t.lower()
        if len(t) < 8 or any(sk in low for sk in _SKIP):
            continue
        if "inut" in low or "4401053694" in t:  # do la ben mua, bo qua
            continue
        if any(kw in low for kw in _ORG_KW):
            # Cat phan nhan neu con dinh (vd "Đơn vị bán hàng: CÔNG TY ...")
            for lbl in ("Đơn vị bán hàng:", "Người bán:", "(Seller):", "Tên đơn vị:"):
                if lbl in t:
                    t = t.split(lbl, 1)[1].strip()
            ten_ban = t
            break
        if not cand:  # du phong: dong "thuc" dau tien
            cand = t
    if not ten_ban:
        ten_ban = cand
    so_hd = _extract_so_hd(raw)
    items = base.get("items", [])
    tong_truoc_thue = tong_thue = tong = 0.0

    # 1) Chinh xac nhat: dong tong hop cuoi hoa don co 3 so lien tiep
    #    (truoc thue, tien thue, tong thanh toan). Thu nhieu mau, KHONG phan biet
    #    hoa/thuong ('Tổng cộng(Total):', 'TỔNG CỘNG (Grand total):', 'Tổng chịu thuế 8%').
    _TRIPLE = r"[^\d]*?([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$"
    for pat in (
        r"tổng cộng\s*\(total\)" + _TRIPLE,
        r"tổng chịu thuế\s*\d{1,2}\s*%[^\n]*?" + _TRIPLE,
        r"(?:tổng cộng|grand total)" + _TRIPLE,
    ):
        m = re.search(pat, raw, re.IGNORECASE | re.MULTILINE)
        if m:
            a, b, c = (money.parse_num(x) for x in m.groups())
            if c >= a > 0 and 0 <= b <= a:  # tong>=truoc thue>0, 0<=thue<=truoc thue
                tong_truoc_thue, tong_thue, tong = a, b, c
                break

    # 2) Fallback: sum cac dong + phat hien thue suat de tinh thue
    if not tong_truoc_thue:
        m = re.search(r"Cộng tiền hàng[^\d]*([\d.,]+)", raw)
        tong_truoc_thue = (
            money.parse_num(m.group(1)) if m
            else sum(money.parse_num(it.get("thanh_tien")) for it in items)
        )
    if not tong:
        mr = re.search(r"[Tt]huế suất[^\d]{0,10}(\d{1,2})\s*%", raw)
        rate = int(mr.group(1)) if mr else 0
        tong_thue = round(tong_truoc_thue * rate / 100)
        tong = round(tong_truoc_thue + tong_thue)
    ngay = _extract_ngay(raw, base.get("ngay")) or ngay_ky  # thieu ngay -> lay ngay ky
    # Thue suat muc HOA DON -> gan cho tung dong (bang PDF khong co cot thue suat).
    inv_rate = _invoice_vat_rate(raw, tong_truoc_thue, tong_thue)
    for it in items:
        it["thue_suat"] = _line_vat(str(it.get("ten") or ""), inv_rate)
    return {
        "source": "pdf",
        "so_hd": so_hd,
        "ky_hieu": _extract_ky_hieu(raw, base.get("ky_hieu", "")),
        "ngay": ngay,
        "ngay_ky": ngay_ky,
        "ten_ban": ten_ban,
        "mst_ban": mst_ban,
        "tong_truoc_thue": tong_truoc_thue,
        "tong_thue": tong_thue,
        "tong_tien": tong,
        "items": items,
        "confidence": 0.8,
        "warnings": [],
    }


_AI_PROMPT = (
    "Bạn nhận văn bản (có thể từ OCR, nhiều lỗi) của MỘT hóa đơn GTGT Việt Nam. "
    "Trích xuất chính xác, KHÔNG bịa. Trả về DUY NHẤT một JSON object, không markdown:\n"
    "{\"so_hd\": \"số hóa đơn\", \"ky_hieu\": \"ký hiệu\", \"ngay\": \"YYYY-MM-DD\", "
    "\"ten_ban\": \"tên đơn vị BÁN\", \"mst_ban\": \"MST bên bán\", "
    "\"tong_truoc_thue\": số, \"tong_thue\": số, \"tong_tien\": số, "
    "\"items\": [{\"ten\": \"\", \"dvt\": \"\", \"so_luong\": số, \"don_gia\": số, "
    "\"thanh_tien\": số, \"thue_suat\": số, \"confidence\": 0..1}], "
    "\"confidence\": 0..1}\n"
    "Trường không đọc được: chuỗi rỗng hoặc 0, và giảm confidence tương ứng."
)


def _pdf_all_text(pdf_bytes: bytes, max_pages: int = 3) -> str:
    """Text cac trang dau; trang nao rong (scan) thi OCR."""
    import pypdfium2 as pdfium

    from . import classify

    d = pdfium.PdfDocument(pdf_bytes)
    parts = []
    for i in range(min(len(d), max_pages)):
        t = d[i].get_textpage().get_text_range() or ""
        parts.append(t)
    text = "\n".join(parts)
    if len(text.strip()) < 50:
        text = classify._ocr_first_page(pdf_bytes)
    return text


def extract_purchase_ai(settings: Settings, pdf_bytes: bytes) -> dict:
    """Trich xuat hoa don scan bang AI. Confidence bi CAP 0.6 -> luon phai duyet tay."""
    text = _pdf_all_text(pdf_bytes)
    if not text.strip():
        raise ai.AIError("Không đọc được chữ nào từ file (OCR thất bại)")
    content = ai.chat(
        settings,
        [
            {"role": "system", "content": _AI_PROMPT},
            {"role": "user", "content": text[:12000]},
        ],
        temperature=0.0,
    )
    try:
        data = ai._parse_json_loose(content)
    except (json.JSONDecodeError, ai.AIError):
        raise ai.AIError(f"AI không trả về JSON hợp lệ: {content[:200]}")
    items = []
    for it in data.get("items") or []:
        items.append({
            "ten": str(it.get("ten") or ""),
            "dvt": str(it.get("dvt") or ""),
            "so_luong": money.parse_num(it.get("so_luong")),
            "don_gia": money.parse_num(it.get("don_gia")),
            "thanh_tien": money.parse_num(it.get("thanh_tien")),
            "thue_suat": money.parse_num(it.get("thue_suat")),
            "confidence": min(0.6, float(it.get("confidence") or 0.5)),
        })
    ngay = str(data.get("ngay") or "")
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", ngay):
        ngay = ""
    return {
        "source": "scan_ai",
        "so_hd": str(data.get("so_hd") or ""),
        "ky_hieu": str(data.get("ky_hieu") or ""),
        "ngay": ngay,
        "ten_ban": str(data.get("ten_ban") or ""),
        "mst_ban": str(data.get("mst_ban") or ""),
        "tong_truoc_thue": money.parse_num(data.get("tong_truoc_thue")),
        "tong_thue": money.parse_num(data.get("tong_thue")),
        "tong_tien": money.parse_num(data.get("tong_tien")),
        "items": items,
        "confidence": min(0.6, float(data.get("confidence") or 0.5)),
        "warnings": [{"code": "ai", "msg": "Trích xuất bằng AI từ bản scan — PHẢI kiểm tra tay từng dòng"}],
    }


# ---------------------------------------------------------------------------
# Match ten hang hoa don -> mat hang ton kho
# ---------------------------------------------------------------------------
def match_suggestions(
    all_items: list[InvItem],
    text: str,
    limit: int = 3,
    aliases: dict[tuple[str, str], tuple[int, int | None]] | None = None,
    mst_ban: str = "",
) -> tuple[InvItem | None, str, list[tuple[InvItem, float]]]:
    """Tra ve (mat hang khop, loai khop, [(goi y, do giong 0..1)]).

    learned: da hoc tu lan gan tay truoc (bang InvItemAlias) — uu tien TRUOC ca
    exact, vi day la lua chon CHINH XAC nguoi dung da xac nhan cho chinh NCC nay.
    exact: trung ten chuan hoa/ma. fuzzy: giong >=85%. Goi y kem diem de UI
    noi ro LY DO (vd 'giống 92% tên').

    aliases: {(ten_norm, mst_ban): (item_id, warehouse_id)} — tra theo (ten, mst_ban
    cua HD hien tai) truoc, roi fallback (ten, "") ap dung chung moi NCC.
    """
    norm = normalize_name(text)
    if not norm:
        return None, "none", []
    if aliases:
        alias = aliases.get((norm, mst_ban)) or aliases.get((norm, ""))
        if alias:
            item_id, _wh_id = alias
            it = next((i for i in all_items if i.id == item_id), None)
            if it is not None:
                return it, "learned", [(it, 1.0)]
    for it in all_items:
        if it.ten_norm == norm or normalize_name(it.ma_hang) == norm:
            return it, "exact", [(it, 1.0)]
    scored = sorted(
        (
            (difflib.SequenceMatcher(None, norm, it.ten_norm).ratio(), it)
            for it in all_items
            if it.ten_norm
        ),
        key=lambda x: -x[0],
    )
    top = [(it, r) for r, it in scored[:limit] if r >= 0.6]
    if top and top[0][1] >= 0.85:
        return top[0][0], "fuzzy", top
    return None, "none", top


def load_purchase_aliases(db: Session) -> dict[tuple[str, str], tuple[int, int | None]]:
    """Tai toan bo alias da hoc 1 lan (dung cho ca request) -> dict cho match_suggestions."""
    return {
        (a.ten_norm, a.mst_ban): (a.item_id, a.warehouse_id)
        for a in db.scalars(select(InvItemAlias))
    }


def create_purchase_draft(
    db: Session,
    data: dict,
    doc_id: str = "",
    doc_suffix: str = ".pdf",
) -> InvPurchase:
    """Tao draft hoa don mua vao tu ket qua parse + doi chieu tong + chong trung."""
    warnings = list(data.get("warnings") or [])
    confidence = float(data.get("confidence") or 0.5)

    # Doi chieu tong tien vs tong cac dong
    sum_lines = sum(money.parse_num(it.get("thanh_tien")) for it in data.get("items") or [])
    tong_truoc_thue = money.parse_num(data.get("tong_truoc_thue"))
    if tong_truoc_thue and abs(sum_lines - tong_truoc_thue) > 1:
        warnings.append({
            "code": "lech_tong",
            "msg": (
                f"Tổng các dòng ({money.vnd(sum_lines)}đ) lệch tổng trước thuế trên "
                f"hóa đơn ({money.vnd(tong_truoc_thue)}đ) — kiểm tra lại từng dòng"
            ),
        })
        confidence = min(confidence, 0.7)
    # Doi chieu cheo: truoc thue + thue phai khop tong thanh toan (phat hien loi
    # OCR/parse nhet nham so, hoac hoa don da bi sua thue suat khong khop)
    tong_thue = money.parse_num(data.get("tong_thue"))
    tong_tien = money.parse_num(data.get("tong_tien"))
    if tong_tien and abs(tong_truoc_thue + tong_thue - tong_tien) > max(1.0, tong_tien * 0.005):
        warnings.append({
            "code": "lech_tong_cong",
            "msg": (
                f"Trước thuế + thuế ({money.vnd(tong_truoc_thue + tong_thue)}đ) lệch tổng "
                f"thanh toán ({money.vnd(tong_tien)}đ) — kiểm tra lại"
            ),
        })
        confidence = min(confidence, 0.7)
    if not data.get("ngay"):
        warnings.append({"code": "thieu_ngay", "msg": "Không đọc được ngày hóa đơn — phải nhập tay"})
        confidence = min(confidence, 0.7)
    # Doi chieu ngay HD vs ngay ky so PDF (chi canh bao khi LECH)
    ngay_ky = str(data.get("ngay_ky") or "")
    if ngay_ky and data.get("ngay") and data["ngay"] != ngay_ky:
        warnings.append({
            "code": "lech_ngay_ky",
            "msg": f"Ngày trên hóa đơn ({data['ngay']}) khác ngày ký số PDF ({ngay_ky}) — kiểm tra lại",
        })

    # Doan loai: NCC dich vu ro rang, HOAC moi dong DVT = 'Gói' (coi la dich vu)
    loai = "hang_hoa"
    ten_ban_norm = normalize_name(str(data.get("ten_ban") or ""))
    items = data.get("items") or []
    item_names = [normalize_name(str(it.get("ten") or "")) for it in items if str(it.get("ten") or "").strip()]
    dvts = [normalize_name(str(it.get("dvt") or "")) for it in items if str(it.get("dvt") or "").strip()]
    all_goi = bool(dvts) and all(d == "goi" for d in dvts)
    # Ten ben ban la dich vu, HOAC MOI dong deu la dich vu (hoa don thuan dich vu).
    # KHONG gan dich vu chi vi 1 dong phi le trong hoa don hang hoa (user tu xoa dong do).
    ten_is_dv = any(kw in ten_ban_norm for kw in _DICH_VU_KEYWORDS)
    all_items_dv = bool(item_names) and all(
        any(kw in nm for kw in _DICH_VU_KEYWORDS) for nm in item_names
    )
    # Co dong ro rang la hang hoa dong goi (vd '500 gr', '1 thung') -> khong doan
    # dich_vu du cac tin hieu khac (ten NCC...) co the trung tu khoa mot cach tinh co.
    has_hang_hoa_unit = any(_HANG_HOA_UNIT_RE.search(nm) for nm in item_names)
    if (all_goi or ten_is_dv or all_items_dv) and not has_hang_hoa_unit:
        loai = "dich_vu"
        warnings.append({
            "code": "doan_dich_vu",
            "msg": "Đoán là hóa đơn DỊCH VỤ/chi phí (không nhập kho) — kiểm tra lại, bấm nút để đổi",
        })

    # Chong trung: (MST ben ban, so HD da chuan hoa — bo so 0 dau)
    dup_of = None
    so_hd_norm = normalize_so_hd(str(data.get("so_hd") or ""))
    if data.get("mst_ban") and so_hd_norm:
        candidates = db.scalars(
            select(InvPurchase).where(InvPurchase.mst_ban == data["mst_ban"])
        )
        dup = next(
            (c for c in candidates if normalize_so_hd(c.so_hd) == so_hd_norm), None
        )
        if dup is not None:
            dup_of = dup.id
            warnings.append({
                "code": "trung_hd",
                "msg": f"Trùng hóa đơn #{dup.id} (cùng MST + số HĐ {data['so_hd']}) — coi chừng import 2 lần",
            })

    inv = InvPurchase(
        so_hd=str(data.get("so_hd") or ""),
        ky_hieu=str(data.get("ky_hieu") or ""),
        mst_ban=str(data.get("mst_ban") or ""),
        ten_ban=str(data.get("ten_ban") or ""),
        ngay=str(data.get("ngay") or ""),
        tong_truoc_thue=tong_truoc_thue,
        tong_thue=tong_thue,
        tong_tien=tong_tien,
        source=str(data.get("source") or "manual"),
        doc_id=doc_id,
        doc_suffix=doc_suffix,
        loai=loai,
        confidence=confidence,
        warnings=json.dumps(warnings, ensure_ascii=False),
        dup_of=dup_of,
    )
    db.add(inv)

    all_items = list(db.scalars(select(InvItem).where(InvItem.active.is_(True))))
    wh_hh = db.scalars(select(InvWarehouse).where(InvWarehouse.code == "HH")).first()
    aliases = load_purchase_aliases(db)
    mst_ban = str(data.get("mst_ban") or "")
    for idx, it in enumerate(data.get("items") or [], start=1):
        ten_raw = str(it.get("ten") or "").strip()
        sl = money.parse_num(it.get("so_luong"))
        dg = money.parse_num(it.get("don_gia"))
        tt = money.parse_num(it.get("thanh_tien"))
        line_warn = []
        line_conf = float(it.get("confidence") or confidence)
        # Suy truong con thieu tu 2 truong con lai — PDF/OCR hay boc thieu 1 cot
        if sl == 0 and dg > 0 and tt > 0:
            q = tt / dg
            if abs(q - round(q)) <= 0.01 * max(q, 1):
                sl = round(q)
                line_warn.append({
                    "code": "suy_so_luong",
                    "msg": f"Suy ra số lượng = {sl:g} từ thành tiền/đơn giá — kiểm tra lại",
                })
                line_conf = min(line_conf, 0.6)
        elif sl > 0 and dg == 0 and tt > 0:
            dg = tt / sl
            line_warn.append({
                "code": "suy_don_gia",
                "msg": f"Suy ra đơn giá = {money.vnd(dg)}đ từ thành tiền/số lượng — kiểm tra lại",
            })
            line_conf = min(line_conf, 0.6)
        elif sl == 0 and dg == 0 and tt > 0:
            sl = 1
            dg = tt
            line_warn.append({
                "code": "suy_dong",
                "msg": f"Thiếu số lượng và đơn giá — tạm coi SL=1, đơn giá={money.vnd(tt)}đ — PHẢI kiểm tra tay",
            })
            line_conf = min(line_conf, 0.6)
        if sl and dg and tt and abs(sl * dg - tt) > max(1.0, tt * 0.01):
            line_warn.append({
                "code": "lech_dong",
                "msg": f"SL×đơn giá ({money.vnd(sl * dg)}) ≠ thành tiền ({money.vnd(tt)})",
            })
            line_conf = min(line_conf, 0.6)
        if not tt and sl and dg:
            tt = round(sl * dg)
        matched, kind, _sugg = match_suggestions(
            all_items, ten_raw, aliases=aliases, mst_ban=mst_ban
        )
        db.add(InvPurchaseLine(
            invoice=inv,
            stt=int(money.parse_num(it.get("stt")) or idx),
            ten_raw=ten_raw,
            dvt=str(it.get("dvt") or ""),
            so_luong=sl,
            don_gia=dg,
            thanh_tien=tt,
            thue_suat=money.parse_num(it.get("thue_suat")),
            item_id=matched.id if matched else None,
            warehouse_id=wh_hh.id if wh_hh else None,
            match_kind=kind,
            confidence=line_conf,
            warnings=json.dumps(line_warn, ensure_ascii=False),
        ))
    db.commit()
    db.refresh(inv)
    return inv


# ---------------------------------------------------------------------------
# Hoa don BAN RA (iNut = ben ban) -> draft. GD1: chi doi chieu, KHONG tru kho.
# ---------------------------------------------------------------------------
INUT_MST = "4401053694"

# ten file: ihoadon_<MST>_<so HD>_<mau>_<ddmmyyyy>_<index>.(pdf|xml)
_FNAME_RE = re.compile(r"ihoadon_\d+_(\d+)_\d+_(\d{2})(\d{2})(\d{4})_", re.IGNORECASE)

_PHAN_MEM_KW = ("phan mem", "license", "software", "ban quyen")
# Bo lap dat / he thong: ban 1 "Bo" nhung ghep tu nhieu linh kien -> can SX (ghep bo)
_ASSEMBLY_KW = (
    "bo thiet bi", "lap dat", "he thong", "tron bo", "combo", "goi thau",
    "lap rap", "thi cong", "trien khai",
)
_CAMERA_KW = (
    "camera", "dau ghi", "o cung", "nvr", "dvr", "the nho", "adapter",
    "nguon", "switch poe", "day mang", "ip",
)


def _is_kct(thue_suat_raw) -> bool:
    """KCT = khong chiu thue (phan mem). TSuat XML cho thang 'KCT'."""
    s = str(thue_suat_raw or "").strip().lower()
    return s in ("", "kct", "khong chiu thue") or "kct" in s


def classify_sale_line(ten: str, thue_kct: bool) -> tuple[str, str]:
    """-> (line_class, fulfil_kind so bo). fulfil se duoc tinh lai o _sale_out theo ton thuc te.

    Thu tu uu tien: phan mem (chua 'license/phan mem') > iNut > camera > khac.
    """
    n = normalize_name(ten)
    if any(k in n for k in _PHAN_MEM_KW):
        return "phan_mem", "doanh_thu"
    if "inut" in n:
        return "inut", "sx"  # mac dinh SX; co ton TP thi _sale_out doi thanh 'ton'
    if any(k in n for k in _ASSEMBLY_KW):
        return "bo", "sx"  # bo lap dat -> ghep bo (khai bao linh kien + AI goi y)
    if any(k in n for k in _CAMERA_KW):
        return "camera", "ton"
    return "other", "none"


def parse_sale_xml(xml_bytes: bytes) -> dict:
    """Parse hoa don BAN RA tu XML: NMua=khach, kiem NBan=iNut, TCHDon=dieu chinh."""
    base = invoice.parse_invoice_xml(xml_bytes)
    root = ET.fromstring(xml_bytes)

    def txt(el, tag, default=""):
        if el is None:
            return default
        x = el.find(tag)
        return (x.text or "").strip() if x is not None and x.text else default

    nban = root.find(".//NBan")
    shd = root.find(".//SHDon")
    tt = root.find(".//TToan")
    mst_ban = txt(nban, "MST")
    tchdon_el = root.find(".//TCHDon")
    tchdon = (tchdon_el.text or "").strip() if tchdon_el is not None else ""

    warnings = []
    if mst_ban and mst_ban != INUT_MST:
        warnings.append({
            "code": "khong_phai_inut",
            "msg": f"Bên bán MST {mst_ban} không phải iNut ({INUT_MST}) — kiểm tra lại file",
        })
    is_dc = tchdon in ("2", "3")
    dc_ref = ""
    if is_dc:
        for it in base.get("items", []):
            if "dieu chinh" in normalize_name(str(it.get("ten") or "")):
                dc_ref = str(it.get("ten"))[:200]
                break

    buyer = base.get("buyer") or {}
    return {
        "source": "xml",
        "so_hd": (shd.text or "").strip() if shd is not None else "",
        "ky_hieu": base.get("ky_hieu", ""),
        "ngay": _ngay_iso(base.get("ngay")),
        "ten_mua": buyer.get("name", ""),
        "mst_mua": buyer.get("mst", ""),
        "tong_truoc_thue": money.parse_num(txt(tt, "TgTCThue")),
        "tong_thue": money.parse_num(txt(tt, "TgTThue")),
        "tong_tien": money.parse_num(txt(tt, "TgTTTBSo")),
        "items": base.get("items", []),
        "is_dieu_chinh": is_dc,
        "dc_ref": dc_ref,
        "confidence": 1.0,
        "warnings": warnings,
    }


def parse_sale_pdf(pdf_bytes: bytes, filename: str = "") -> dict:
    """Parse hoa don BAN RA tu PDF (fallback khi khong co XML). Kem chinh xac hon."""
    base = invoice.parse_invoice(pdf_bytes)
    raw = base.get("raw_text", "")
    ngay_ky = _pdf_sign_date(pdf_bytes)
    buyer = base.get("buyer") or {}

    warnings = []
    if INUT_MST not in raw and "inut" not in raw.lower():
        warnings.append({
            "code": "khong_phai_inut",
            "msg": "Không thấy iNut là bên bán trong file — kiểm tra lại",
        })
    is_dc = "dieu chinh cho hoa don" in normalize_name(raw)
    dc_ref = ""
    if is_dc:
        m = re.search(r"(Điều chỉnh[^\n]{0,90})", raw, re.IGNORECASE)
        dc_ref = m.group(1).strip()[:200] if m else "điều chỉnh"

    so_hd = _extract_so_hd(raw)
    ngay = _extract_ngay(raw, base.get("ngay")) or ngay_ky
    fm = _FNAME_RE.search(filename or "")
    if fm:  # fallback tu ten file
        if not so_hd:
            so_hd = fm.group(1)
        if not ngay:
            ngay = _mk_iso(fm.group(4), fm.group(3), fm.group(2))

    items = base.get("items", [])
    tong_truoc = sum(money.parse_num(it.get("thanh_tien")) for it in items)
    m = re.search(r"Tổng cộng tiền thanh toán[^\d]*([\d.,]+)", raw)
    tong_tien = money.parse_num(m.group(1)) if m else tong_truoc
    tong_thue = max(0.0, round(tong_tien - tong_truoc))
    return {
        "source": "pdf",
        "so_hd": so_hd,
        "ky_hieu": _extract_ky_hieu(raw, base.get("ky_hieu", "")),
        "ngay": ngay,
        "ngay_ky": ngay_ky,
        "ten_mua": buyer.get("name", ""),
        "mst_mua": buyer.get("mst", ""),
        "tong_truoc_thue": tong_truoc,
        "tong_thue": tong_thue,
        "tong_tien": tong_tien,
        "items": items,
        "is_dieu_chinh": is_dc,
        "dc_ref": dc_ref,
        "confidence": 0.75,
        "warnings": warnings,
    }


def create_sale_draft(
    db: Session, data: dict, doc_id: str = "", doc_suffix: str = ".pdf"
) -> InvSale:
    """Tao draft hoa don ban ra + match + phan loai dong. KHONG tao InvMove (GD1)."""
    warnings = list(data.get("warnings") or [])
    confidence = float(data.get("confidence") or 0.6)
    is_dc = bool(data.get("is_dieu_chinh"))

    sum_lines = sum(money.parse_num(it.get("thanh_tien")) for it in data.get("items") or [])
    tong_truoc = money.parse_num(data.get("tong_truoc_thue"))
    if not is_dc and tong_truoc and abs(sum_lines - tong_truoc) > 1:
        warnings.append({
            "code": "lech_tong",
            "msg": f"Tổng dòng ({money.vnd(sum_lines)}đ) lệch tổng trước thuế ({money.vnd(tong_truoc)}đ) — soát lại",
        })
        confidence = min(confidence, 0.7)
    if not data.get("ngay"):
        warnings.append({"code": "thieu_ngay", "msg": "Không đọc được ngày hóa đơn — nhập tay"})
    if is_dc:
        warnings.append({
            "code": "dieu_chinh",
            "msg": "Hóa đơn điều chỉnh/thay thế — bỏ qua đối chiếu kho, chỉ lưu vết",
        })

    # Chong trung: (ky hieu + so HD chuan hoa) — iNut la ben ban
    dup_of = None
    so_hd_norm = normalize_so_hd(str(data.get("so_hd") or ""))
    ky_hieu = str(data.get("ky_hieu") or "")
    if so_hd_norm and ky_hieu:
        for c in db.scalars(select(InvSale).where(InvSale.ky_hieu == ky_hieu)):
            if normalize_so_hd(c.so_hd) == so_hd_norm:
                dup_of = c.id
                warnings.append({
                    "code": "trung_hd",
                    "msg": f"Trùng HĐ bán #{c.id} (cùng ký hiệu + số {data.get('so_hd')}) — coi chừng import 2 lần",
                })
                break

    inv = InvSale(
        so_hd=str(data.get("so_hd") or ""),
        ky_hieu=ky_hieu,
        mst_mua=str(data.get("mst_mua") or ""),
        ten_mua=str(data.get("ten_mua") or ""),
        ngay=str(data.get("ngay") or ""),
        tong_truoc_thue=tong_truoc,
        tong_thue=money.parse_num(data.get("tong_thue")),
        tong_tien=money.parse_num(data.get("tong_tien")),
        source=str(data.get("source") or "manual"),
        doc_id=doc_id,
        doc_suffix=doc_suffix,
        is_dieu_chinh=is_dc,
        dc_ref=str(data.get("dc_ref") or "")[:255],
        confidence=confidence,
        warnings=json.dumps(warnings, ensure_ascii=False),
        dup_of=dup_of,
    )
    # Tu gan khach hang neu MST/ten khop san (khong tao moi tu luong nay)
    matched = accounts.find_customer(db, inv.ten_mua, inv.mst_mua)
    if matched:
        inv.customer_id = matched.id
    db.add(inv)

    all_items = list(db.scalars(select(InvItem).where(InvItem.active.is_(True))))
    wh_tp = db.scalars(select(InvWarehouse).where(InvWarehouse.code == "TP")).first()
    wh_hh = db.scalars(select(InvWarehouse).where(InvWarehouse.code == "HH")).first()
    for idx, it in enumerate(data.get("items") or [], start=1):
        ten_raw = str(it.get("ten") or "").strip()
        sl = money.parse_num(it.get("so_luong"))
        dg = money.parse_num(it.get("don_gia"))
        tt = money.parse_num(it.get("thanh_tien"))
        if not tt and sl and dg:
            tt = round(sl * dg)
        ts_raw = it.get("thue_suat")
        kct = _is_kct(ts_raw)
        if is_dc:
            line_class, fulfil = "other", "none"
        else:
            line_class, fulfil = classify_sale_line(ten_raw, kct)
        matched, kind = None, "none"
        if not is_dc and fulfil != "doanh_thu":
            matched, kind, _sugg = match_suggestions(all_items, ten_raw)
        wh = wh_tp if line_class == "inut" else wh_hh
        db.add(InvSaleLine(
            invoice=inv,
            stt=int(money.parse_num(it.get("stt")) or idx),
            ten_raw=ten_raw,
            dvt=str(it.get("dvt") or ""),
            so_luong=sl,
            don_gia_ban=dg,
            thanh_tien=tt,
            thue_suat=money.parse_num(ts_raw),
            thue_kct=kct,
            item_id=matched.id if matched else None,
            warehouse_id=(wh.id if wh else None),
            match_kind=kind,
            line_class=line_class,
            fulfil_kind=fulfil,
            confidence=confidence,
        ))
    db.commit()
    db.refresh(inv)
    return inv


# ---------------------------------------------------------------------------
# Doi chieu bang ke hoa don mua vao (Excel ke khai thue, sheet "Bảng kê mua vào")
# ---------------------------------------------------------------------------
def parse_bang_ke_xlsx(data: bytes) -> list[dict]:
    """Doc sheet 'Bảng kê mua vào' (phu luc to khai thue GTGT) -> danh sach dong HD.

    Cot: A=STT, B=so HD, C=ngay lap (dd/mm/yyyy), D=ten nguoi ban,
    E=gia tri truoc thue, G=thue GTGT.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "mua vao" in normalize_name(name):
            ws = wb[name]
            break
    if ws is None:
        raise ValueError("Không tìm thấy sheet 'Bảng kê mua vào' trong file")

    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        a_s = str(a).strip() if a is not None else ""
        b_s = str(b).strip() if b is not None else ""
        if not a_s.isdigit() or not b_s:
            continue
        c = row[2] if len(row) > 2 else None
        ngay = ""
        if isinstance(c, datetime):
            ngay = c.strftime("%Y-%m-%d")
        else:
            m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(c or "").strip())
            if m:
                ngay = _mk_iso(m.group(3), m.group(2), m.group(1))
        ten_ban = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        gia_tri = money.parse_num(row[4]) if len(row) > 4 else 0.0
        thue = money.parse_num(row[6]) if len(row) > 6 else 0.0
        rows.append({
            "so_hd": b_s,
            "ngay": ngay,
            "ten_ban": ten_ban,
            "gia_tri": gia_tri,
            "thue": thue,
        })
    if not rows:
        raise ValueError("Không đọc được dòng nào trong bảng kê mua vào — kiểm tra lại file")
    return rows


def reconcile_bang_ke(db: Session, rows: list[dict]) -> dict:
    """Doi chieu bang ke thue voi cac hoa don da import.

    khop: so HD + tien khop. lech_tien: so HD khop, tien lech >2đ. thieu_file: khong
    tim thay hoa don. ngoai_bang_ke: hoa don (draft/posted) khong xuat hien trong bang ke.
    """
    purchases = list(db.scalars(select(InvPurchase).where(InvPurchase.status != "void")))
    by_norm: dict[str, list[InvPurchase]] = {}
    for p in purchases:
        by_norm.setdefault(normalize_so_hd(p.so_hd), []).append(p)

    matched_ids: set[int] = set()
    khop: list[dict] = []
    lech_tien: list[dict] = []
    thieu_file: list[dict] = []

    for r in rows:
        norm = normalize_so_hd(r["so_hd"])
        # So HD la khoa chinh: khop so la nhan, roi phan loai khop/lech tien theo
        # so tien. Uu tien ung vien khop tien; nhieu ung vien thi lay ten gan nhat.
        cands = [p for p in by_norm.get(norm, []) if p.id not in matched_ids]
        found = None
        for p in cands:
            if abs((p.tong_truoc_thue or 0.0) - r["gia_tri"]) <= 2:
                found = p
                break
        if found is None and cands:
            found = max(
                cands,
                key=lambda p: difflib.SequenceMatcher(
                    None, normalize_name(p.ten_ban), normalize_name(r["ten_ban"])
                ).ratio(),
            )
        if found is None:
            thieu_file.append({
                "so_hd": r["so_hd"], "ngay": r["ngay"],
                "ten_ban": r["ten_ban"], "gia_tri": r["gia_tri"],
            })
            continue

        matched_ids.add(found.id)
        if found.status == "draft" and not found.ngay and r["ngay"]:
            found.ngay = r["ngay"]
            warns = json.loads(found.warnings or "[]")
            warns.append({"code": "ngay_bang_ke", "msg": "Ngày lấy từ bảng kê thuế"})
            found.warnings = json.dumps(warns, ensure_ascii=False)

        entry = {
            "so_hd": r["so_hd"], "ngay": r["ngay"], "ten_ban": r["ten_ban"],
            "gia_tri": r["gia_tri"], "purchase_id": found.id,
        }
        if abs((found.tong_truoc_thue or 0.0) - r["gia_tri"]) > 2:
            entry["purchase_gia_tri"] = found.tong_truoc_thue
            lech_tien.append(entry)
        else:
            khop.append(entry)

    ngoai_bang_ke = [
        {
            "so_hd": p.so_hd, "ngay": p.ngay, "ten_ban": p.ten_ban,
            "gia_tri": p.tong_truoc_thue, "purchase_id": p.id,
        }
        for p in purchases
        if p.id not in matched_ids
    ]

    db.commit()
    return {
        "khop": khop,
        "lech_tien": lech_tien,
        "thieu_file": thieu_file,
        "ngoai_bang_ke": ngoai_bang_ke,
    }


# ---------------------------------------------------------------------------
# Import tu URL (Google Drive hoac link file truc tiep) + giai nen ZIP
# ---------------------------------------------------------------------------
def fetch_from_url(url: str, max_files: int = 50) -> list[tuple[str, bytes]]:
    """Tai file/thu muc tu URL -> [(ten file, noi dung)]. Ho tro Google Drive + link file truc tiep."""
    url = (url or "").strip()
    if not url:
        raise ValueError("Thiếu URL")

    if "drive.google.com" in url:
        import gdown

        with tempfile.TemporaryDirectory() as tmpdir:
            try:
                if "/folders/" in url:
                    gdown.download_folder(url, output=tmpdir, quiet=True, use_cookies=False)
                else:
                    gdown.download(url=url, output=tmpdir + "/", fuzzy=True, quiet=True)
            except Exception as e:  # noqa: BLE001
                raise ValueError(f"Không tải được từ Google Drive: {e}")
            out: list[tuple[str, bytes]] = []
            for p in sorted(Path(tmpdir).rglob("*")):
                if not p.is_file() or p.suffix.lower() not in (".pdf", ".xml", ".zip"):
                    continue
                out.append((p.name, p.read_bytes()))
                if len(out) >= max_files:
                    break
            if not out:
                raise ValueError("Không tìm thấy file PDF/XML/ZIP nào trong link Google Drive")
            return out

    try:
        r = httpx.get(url, timeout=60, follow_redirects=True)
        r.raise_for_status()
    except httpx.HTTPError as e:
        raise ValueError(f"Không tải được file từ URL: {e}")
    name = unquote(Path(urlparse(url).path).name) or "downloaded"
    return [(name, r.content)]


_MAIN_XML_MARK = b"<HDon"


def expand_zip(name: str, data: bytes) -> list[tuple[str, bytes]]:
    """Giai nen ZIP: neu co XML hoa don chinh (chua '<HDon', bo *_Bang_Ke*/*hdcd*)
    -> chi lay cac XML do; khong thi lay cac PDF. Khong phai ZIP -> tra ve nguyen ven."""
    if not (name.lower().endswith(".zip") or data[:2] == b"PK"):
        return [(name, data)]

    xmls: list[tuple[str, bytes]] = []
    pdfs: list[tuple[str, bytes]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                fname = Path(info.filename).name
                low = fname.lower()
                if low.endswith(".xml"):
                    xmls.append((fname, zf.read(info)))
                elif low.endswith(".pdf"):
                    pdfs.append((fname, zf.read(info)))
    except zipfile.BadZipFile:
        return [(name, data)]

    main_xmls = [
        (n, c) for n, c in xmls
        if _MAIN_XML_MARK in c and "bang_ke" not in n.lower() and "hdcd" not in n.lower()
    ]
    if main_xmls:
        return main_xmls
    return pdfs
