"""Parse to khai nhap khau VNACCS (Excel 7N) + giay nop tien thue + dien MT103.

3 parser thuan (nhan bytes, tra dict, khong dung DB) — dung o lop tren
(inv_api.py) de tao InvCustomsDecl/Line/Cost.
"""
from __future__ import annotations

import io
import re

from . import money
from .inv_import import _pdf_all_text
from .inventory import normalize_name

_LOAI_HINH_RE = re.compile(r"^([A-Z]\d{2})")
_MARKER_RE = re.compile(r"^<(\d{2})>$")
_WINDOW = 35  # so dong quet sau moi marker de tim nhan (label) - chiu xe dich


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def _norm_dvt(v: str) -> str:
    # Giu ma phap ly tren to khai (PCE, KGM...) de doi chieu khong bi sai khac.
    return v.strip().upper()


def _pct(v) -> float:
    """'8%' -> 8.0 ; '0%' -> 0.0."""
    return money.parse_num(_s(v).replace("%", ""))


def _mk_iso_slash(s: str) -> str:
    """'28/05/2026' -> '2026-05-28'. Rong/khong khop -> ''."""
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", (s or "").strip())
    if not m:
        return ""
    d, mo, y = m.groups()
    return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"


# ---------------------------------------------------------------------------
# Excel to khai VNACCS 7N (sheet "TKN")
# ---------------------------------------------------------------------------
def _row_map(ws, r: int) -> dict[str, object]:
    return {c.column_letter: c.value for c in ws[r] if c.value is not None}


def _find_row(
    rows: list[tuple[int, dict]], label: str, after: int = 0
) -> tuple[int, dict] | None:
    """Tim dong dau tien (sau dong `after`) co 1 o BAT DAU bang `label` (chuan hoa).

    Dung startswith (khong phai substring 'in') de tranh khop nham: vd nhan
    'Mã áp dụng thuế suất' chua chuoi con 'thuế suất' nhung KHONG phai nhan
    'Thuế suất' (2 nhan khac nhau, gia tri o cot khac nhau).
    """
    label_n = normalize_name(label)
    for r, cells in rows:
        if r <= after:
            continue
        for v in cells.values():
            if normalize_name(str(v)).startswith(label_n):
                return r, cells
    return None


def parse_customs_xlsx(data: bytes) -> dict:
    """Doc file to khai nhap khau VNACCS 7N -> header + danh sach dong hang.

    Sheet "TKN": header 1 lan (E4 so to khai...), dong hang danh dau bang
    marker '<01>', '<02>'... o mot o rieng (thuong cot C). Tu marker, quet
    ~35 dong ke tiep TIM NHAN (label) roi doc gia tri o cung dong — KHONG
    hard-code khoang cach tuyet doi, vi bang co the lech dong giua cac file.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Không đọc được file Excel tờ khai: {e}")

    if "TKN" not in wb.sheetnames:
        raise ValueError("File không đúng định dạng tờ khai VNACCS 7N — thiếu sheet 'TKN'")
    ws = wb["TKN"]

    so_to_khai = _s(ws["E4"].value)
    if not so_to_khai or not so_to_khai.isdigit():
        raise ValueError("Không đọc được số tờ khai (ô E4) — kiểm tra lại file có đúng tờ khai VNACCS không")

    ma_loai_hinh_raw = _s(ws["P6"].value)
    m = _LOAI_HINH_RE.match(ma_loai_hinh_raw)
    ma_loai_hinh = m.group(1) if m else ma_loai_hinh_raw[:5]

    ngay_dang_ky = _mk_iso_slash(_s(ws["G8"].value).split(" ")[0])

    incoterm_ccy = _s(ws["J45"].value)  # "A - DAP - USD - "
    parts = [p.strip() for p in incoterm_ccy.split("-") if p.strip()]
    incoterm = parts[1] if len(parts) > 1 else ""
    nguyen_te = parts[2] if len(parts) > 2 else ""

    # --- Danh sach dong hang: quet toan bo sheet tim marker <NN> ---
    max_row = ws.max_row
    markers: list[tuple[int, int]] = []  # (stt, row)
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if cell.value is None:
                continue
            mm = _MARKER_RE.match(str(cell.value).strip())
            if mm:
                markers.append((int(mm.group(1)), cell.row))
                break
    if not markers:
        raise ValueError("Không tìm thấy dòng hàng nào (marker <NN>) trong tờ khai")

    lines: list[dict] = []
    for idx, (stt, r0) in enumerate(markers):
        r_end = markers[idx + 1][1] if idx + 1 < len(markers) else min(r0 + _WINDOW, max_row + 1)
        r_end = min(r_end, r0 + _WINDOW, max_row + 1)
        window = [(r, _row_map(ws, r)) for r in range(r0, r_end)]

        res = _find_row(window, "Mã số hàng hóa")
        ma_hs = _s(res[1].get("G")) if res else ""

        res = _find_row(window, "Mô tả hàng hóa")
        mo_ta = _s(res[1].get("G")) if res else ""

        res = _find_row(window, "Số lượng (1)")
        so_luong = money.parse_num(res[1].get("V")) if res else 0.0
        dvt = _norm_dvt(_s(res[1].get("AE"))) if res else ""

        res = _find_row(window, "Trị giá hóa đơn")
        tri_gia_nt = money.parse_num(res[1].get("I")) if res else 0.0
        don_gia_nt = money.parse_num(res[1].get("V")) if res else 0.0

        res_thue = _find_row(window, "Trị giá tính thuế")
        tri_gia_tinh_thue = money.parse_num(res_thue[1].get("I")) if res_thue else 0.0
        divider_row = res_thue[0] if res_thue else r0
        after_thue = divider_row

        res = _find_row(window, "Thuế suất", after=after_thue - 1)
        thue_suat_nk = _pct(res[1].get("I")) if res else 0.0

        res = _find_row(window, "Số tiền thuế", after=after_thue - 1)
        tien_thue_nk = money.parse_num(res[1].get("I")) if res else 0.0

        res = _find_row(window, "Nước xuất xứ")
        nuoc_xx = _s(res[1].get("X")) if res else ""

        res_div = _find_row(window, "Thuế và thu khác")
        divider2 = res_div[0] if res_div else r0

        res = _find_row(window, "Thuế suất", after=divider2)
        thue_suat_vat = _pct(res[1].get("I")) if res else 0.0

        res = _find_row(window, "Số tiền thuế", after=divider2)
        tien_thue_vat = money.parse_num(res[1].get("I")) if res else 0.0

        lines.append({
            "stt": stt,
            "ma_hs": ma_hs,
            "mo_ta": mo_ta,
            "so_luong": so_luong,
            "dvt": dvt,
            "don_gia_nt": don_gia_nt,
            "tri_gia_nt": tri_gia_nt,
            "tri_gia_tinh_thue": tri_gia_tinh_thue,
            "thue_suat_nk": thue_suat_nk,
            "tien_thue_nk": tien_thue_nk,
            "thue_suat_vat": thue_suat_vat,
            "tien_thue_vat": tien_thue_vat,
            "nuoc_xk": nuoc_xx,
        })

    return {
        "so_to_khai": so_to_khai,
        "phan_luong": _s(ws["I6"].value),
        "ma_loai_hinh": ma_loai_hinh,
        "co_quan_hq": _s(ws["L7"].value),
        "ngay_dang_ky": ngay_dang_ky,
        "nguoi_xk": _s(ws["H23"].value),
        "nuoc_xk": _s(ws["H27"].value),
        "so_van_don": _s(ws["D31"].value),
        "so_hoa_don": _s(ws["J41"].value),
        "ngay_hoa_don": _mk_iso_slash(_s(ws["J43"].value)),
        "phuong_thuc_tt": _s(ws["J44"].value),
        "incoterm": incoterm,
        "nguyen_te": nguyen_te,
        "tri_gia_nt": money.parse_num(ws["P45"].value),
        "phi_ship_nt": money.parse_num(ws["L55"].value),
        "ti_gia": money.parse_num(ws["AB70"].value),
        "tri_gia_tinh_thue": money.parse_num(ws["J46"].value),
        "tong_thue_nk": round(sum(ln["tien_thue_nk"] for ln in lines), 2),
        "tong_thue_vat": money.parse_num(ws["H68"].value),
        "lines": lines,
    }


# ---------------------------------------------------------------------------
# PDF giay nop tien vao NSNN (mau C1-02/NS)
# ---------------------------------------------------------------------------
_ROW_START_RE = re.compile(r"\b(\d{1,2})\s+(\d{9,12})\s+(\d{1,2}/\d{1,2}/\d{4})\s+")

_NDKT_MAP = {
    "1901": "thue_nk",
    "1702": "vat",
    "1751": "ttdb",
}


def _classify_ndkt(ma_ndkt: str, noi_dung: str) -> str:
    if ma_ndkt in _NDKT_MAP:
        return _NDKT_MAP[ma_ndkt]
    n = normalize_name(noi_dung)
    if "gia tri gia tang" in n:
        return "vat"
    if "tieu thu dac biet" in n:
        return "ttdb"
    if "nhap khau" in n and "thue" in n:
        return "thue_nk"
    return "le_phi"


def parse_giay_nop_tien(pdf_bytes: bytes) -> dict:
    """Doc giay nop tien vao NSNN (mau C1-02/NS) -> so to khai (prefix) + cac khoan nop.

    Bang co the bi ngat dong giua so (vd nam '2026' bi tach '202'+'6' o dong sau) do
    PDF wrap chu — noi lai cac chu so lien tiep bi ngat dong truoc khi doc.
    """
    raw = _pdf_all_text(pdf_bytes).replace("\r\n", "\n").replace("\r", "\n")
    if "NGÂN SÁCH NHÀ NƯỚC" not in raw.upper():
        raise ValueError("Không phải giấy nộp tiền vào NSNN (mẫu C1-02/NS)")

    norm = re.sub(r"(?<=\d)\n(?=\d)", "", raw)
    norm = re.sub(r"\s+", " ", norm)

    starts = list(_ROW_START_RE.finditer(norm))
    if not starts:
        raise ValueError("Không đọc được dòng khoản nộp trong giấy nộp tiền")

    end_tong = norm.find("Tổng cộng")
    khoan_nop: list[dict] = []
    so_to_khai_prefix = ""
    ngay = ""
    for i, m in enumerate(starts):
        _stt, so_tk, ngay_m = m.groups()
        seg_start = m.end()
        seg_end = starts[i + 1].start() if i + 1 < len(starts) else (
            end_tong if end_tong != -1 else len(norm)
        )
        tokens = norm[seg_start:seg_end].strip().split()
        if len(tokens) < 4:
            continue
        ndkt, chuong = tokens[-2], tokens[-3]  # tokens[-1] la ma DBHC — khong dung toi
        rest = tokens[:-3]
        j = len(rest)
        while j > 0 and rest[j - 1].isdigit():
            j -= 1
        noi_dung = " ".join(rest[:j]).strip()
        amount_tokens = rest[j:]
        half = len(amount_tokens) // 2 or len(amount_tokens)
        so_tien = money.parse_num("".join(amount_tokens[half:])) if amount_tokens else 0.0

        if not so_to_khai_prefix:
            so_to_khai_prefix = so_tk
            ngay = _mk_iso_slash(ngay_m)

        khoan_nop.append({
            "noi_dung": noi_dung,
            "so_tien": so_tien,
            "ma_chuong": chuong,
            "ma_ndkt": ndkt,
            "phan_loai": _classify_ndkt(ndkt, noi_dung),
        })

    if not khoan_nop:
        raise ValueError("Không đọc được khoản nộp nào trong giấy nộp tiền")

    return {"so_to_khai_prefix": so_to_khai_prefix, "ngay": ngay, "khoan_nop": khoan_nop}


# ---------------------------------------------------------------------------
# Dien MT103 (chuyen tien quoc te) — goi y phi ngan hang
# ---------------------------------------------------------------------------
_MT103_32A_RE = re.compile(r"32A:[^\n]*\n(\d{2})(\d{2})(\d{2})([A-Z]{3})([\d.,]+)")
_MT103_70_RE = re.compile(r"70:[^\n]*\n([^\n]+)")
_MT103_59_RE = re.compile(r"59:[^\n]*\n(?:/\S+\n)?([^\n]+)")


def parse_mt103(pdf_bytes: bytes) -> dict:
    """Doc dien MT103 -> ngay/nguyen te/so tien/noi dung chuyen tien (goi y phi ngan hang)."""
    raw = _pdf_all_text(pdf_bytes).replace("\r\n", "\n").replace("\r", "\n")

    m = _MT103_32A_RE.search(raw)
    if not m:
        raise ValueError("Không phải điện MT103 (thiếu trường 32A: Value Date/Currency/Amount)")
    yy, mo, dd, ccy, amt = m.groups()
    ngay = f"20{yy}-{mo}-{dd}"
    so_tien_nt = money.parse_num(amt)

    m70 = _MT103_70_RE.search(raw)
    remittance = m70.group(1).strip() if m70 else ""

    m59 = _MT103_59_RE.search(raw)
    beneficiary = m59.group(1).strip() if m59 else ""

    return {
        "ngay": ngay,
        "nguyen_te": ccy,
        "so_tien_nt": so_tien_nt,
        "remittance": remittance,
        "beneficiary": beneficiary,
    }
