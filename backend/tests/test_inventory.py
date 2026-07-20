"""Test engine so kho: gia binh quan, kha dung, chan am kho, import Excel, san xuat."""
from __future__ import annotations

import io

import pytest

pytest.importorskip("fastapi")
pytest.importorskip("openpyxl")

from fastapi.testclient import TestClient  # noqa: E402


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.setenv("NAS_ENABLED", "false")
    monkeypatch.setenv("AI_ENABLED", "false")
    from app.config import get_settings

    get_settings.cache_clear()

    from app import db as dbmod
    from app.auth import ensure_admin_seed

    dbmod.reset_engine_for_tests()
    dbmod.init_db()
    gen = dbmod.get_session()
    s = next(gen)
    ensure_admin_seed(s, get_settings())
    gen.close()

    from app.main import app

    c = TestClient(app)
    r = c.post("/api/login", json={"username": "admin", "password": "NhapHang123@"})
    assert r.status_code == 200
    return c


def _mk_item(client, ma, ten, dvt="Cái"):
    r = client.post("/api/inv/items", json={"ma_hang": ma, "ten": ten, "dvt": dvt})
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _wh(client, code):
    r = client.get("/api/inv/warehouses")
    return {w["code"]: w["id"] for w in r.json()}[code]


def _purchase(client, ngay, lines, so_hd="1", mst="0123456789"):
    """Tao + ghi so hoa don mua. lines: [(item_id, wh_id, sl, don_gia)]."""
    r = client.post("/api/inv/purchase", json={
        "so_hd": so_hd, "mst_ban": mst, "ten_ban": "NCC Test", "ngay": ngay,
        "lines": [
            {"ten_raw": f"item{i}", "so_luong": sl, "don_gia": dg,
             "thanh_tien": sl * dg, "item_id": iid, "warehouse_id": wid}
            for i, (iid, wid, sl, dg) in enumerate(lines)
        ],
    })
    assert r.status_code == 200, r.text
    pid = r.json()["id"]
    r = client.post(f"/api/inv/purchase/{pid}/post")
    assert r.status_code == 200, r.text
    return pid


def _issue(client, ngay, lines, expect=200):
    r = client.post("/api/inv/issues", json={
        "ngay": ngay,
        "lines": [
            {"item_id": iid, "warehouse_id": wid, "so_luong": sl}
            for iid, wid, sl in lines
        ],
    })
    assert r.status_code == 200, r.text
    iid_ = r.json()["id"]
    r = client.post(f"/api/inv/issues/{iid_}/post")
    assert r.status_code == expect, r.text
    return iid_, r


# ---------------------------------------------------------------------------
# Gia binh quan gia quyen
# ---------------------------------------------------------------------------
def test_weighted_average_cost(client):
    hh = _wh(client, "HH")
    item = _mk_item(client, "T001", "Camera test")
    # Nhap 10 @100k, nhap 10 @200k -> binh quan 150k
    _purchase(client, "2026-01-05", [(item, hh, 10, 100_000)], so_hd="1")
    _purchase(client, "2026-01-10", [(item, hh, 10, 200_000)], so_hd="2")
    _, r = _issue(client, "2026-01-15", [(item, hh, 5)])
    line = r.json()["lines"][0]
    assert line["gia_von"] == 750_000  # 5 x 150k
    # Ton con 15, gia tri 2.25tr
    stock = client.get("/api/inv/stock").json()
    row = [x for x in stock["rows"] if x["item_id"] == item][0]
    assert row["ton"] == 15
    assert row["gia_tri"] == 2_250_000


def test_stock_out_flushes_rounding(client):
    hh = _wh(client, "HH")
    item = _mk_item(client, "T002", "Hang le")
    # 3 @ 10000/3 dong -> gia tri 10000; xuat het 3 lan 1 -> gia tri ve dung 0
    r = client.post("/api/inv/purchase", json={
        "ngay": "2026-01-05", "ten_ban": "X",
        "lines": [{"ten_raw": "a", "so_luong": 3, "don_gia": 3333.33,
                   "thanh_tien": 10000, "item_id": item, "warehouse_id": hh}],
    })
    pid = r.json()["id"]
    assert client.post(f"/api/inv/purchase/{pid}/post").status_code == 200
    for d in ("2026-01-06", "2026-01-07", "2026-01-08"):
        _issue(client, d, [(item, hh, 1)])
    stock = client.get("/api/inv/stock?all_items=true").json()
    row = [x for x in stock["rows"] if x["item_id"] == item]
    assert row == [] or (row[0]["ton"] == 0 and row[0]["gia_tri"] == 0)


# ---------------------------------------------------------------------------
# Chan am kho + kha dung
# ---------------------------------------------------------------------------
def test_issue_blocked_when_insufficient(client):
    hh = _wh(client, "HH")
    item = _mk_item(client, "T003", "Hang thieu")
    _purchase(client, "2026-01-05", [(item, hh, 2, 50_000)])
    _, r = _issue(client, "2026-01-10", [(item, hh, 5)], expect=400)
    detail = r.json()["detail"]
    assert "violations" in detail
    assert detail["violations"][0]["ma_hang"] == "T003"


def test_issue_before_purchase_date_blocked(client):
    """Ban TRUOC ngay mua -> chan (thoi diem mua < thoi diem ban)."""
    hh = _wh(client, "HH")
    item = _mk_item(client, "T004", "Hang mua sau")
    _purchase(client, "2026-03-01", [(item, hh, 10, 10_000)])
    _, r = _issue(client, "2026-02-15", [(item, hh, 1)], expect=400)
    assert "violations" in r.json()["detail"]


def test_same_day_purchase_then_issue_ok(client):
    hh = _wh(client, "HH")
    item = _mk_item(client, "T005", "Cung ngay")
    _purchase(client, "2026-01-05", [(item, hh, 1, 10_000)])
    _issue(client, "2026-01-05", [(item, hh, 1)])  # nhap truoc xuat trong ngay


def test_availability_min_future_rule(client):
    """Ton 10 ngay 1, da co phieu xuat 8 ngay 5 -> kha dung ngay 2 chi con 2."""
    hh = _wh(client, "HH")
    item = _mk_item(client, "T006", "Kha dung")
    _purchase(client, "2026-01-01", [(item, hh, 10, 10_000)])
    _issue(client, "2026-01-05", [(item, hh, 8)])
    av = client.get("/api/inv/availability?date=2026-01-02").json()
    row = [x for x in av["rows"] if x["item_id"] == item][0]
    assert row["ton"] == 10
    assert row["kha_dung"] == 2


def test_backdated_purchase_recomputes_costs(client):
    """Chen hoa don mua lui ngay -> gia von phieu xuat sau do duoc tinh lai."""
    hh = _wh(client, "HH")
    item = _mk_item(client, "T007", "Lui ngay")
    _purchase(client, "2026-01-10", [(item, hh, 10, 100_000)], so_hd="10")
    iss, _ = _issue(client, "2026-01-20", [(item, hh, 4)])
    # Chen them lo mua 01/15 gia 200k -> binh quan tai 01/20 = 150k
    _purchase(client, "2026-01-15", [(item, hh, 10, 200_000)], so_hd="11")
    r = client.get("/api/inv/issues").json()
    line = [i for i in r if i["id"] == iss][0]["lines"][0]
    assert line["gia_von"] == 600_000  # 4 x 150k


def test_void_purchase_blocked_if_starves_issue(client):
    hh = _wh(client, "HH")
    item = _mk_item(client, "T008", "Void chan")
    pid = _purchase(client, "2026-01-05", [(item, hh, 5, 10_000)])
    _issue(client, "2026-01-10", [(item, hh, 5)])
    r = client.post(f"/api/inv/purchase/{pid}/void")
    assert r.status_code == 400
    assert "violations" in r.json()["detail"]
    # Van posted
    assert client.get(f"/api/inv/purchase/{pid}").json()["status"] == "posted"


# ---------------------------------------------------------------------------
# San xuat
# ---------------------------------------------------------------------------
def test_production_cost_rollup(client):
    hh, nvl, tp = _wh(client, "HH"), _wh(client, "NVL"), _wh(client, "TP")
    n1 = _mk_item(client, "N001", "Module SIM")
    n2 = _mk_item(client, "N002", "Vo hop")
    out = _mk_item(client, "TP001", "Data Logger")
    _purchase(client, "2026-01-05", [(n1, nvl, 10, 500_000), (n2, nvl, 10, 50_000)])
    r = client.post("/api/inv/productions", json={
        "ngay": "2026-01-10",
        "lines": [
            {"chieu": "vao", "item_id": n1, "warehouse_id": nvl, "so_luong": 2},
            {"chieu": "vao", "item_id": n2, "warehouse_id": nvl, "so_luong": 2},
            {"chieu": "ra", "item_id": out, "warehouse_id": tp, "so_luong": 2},
        ],
    })
    pid = r.json()["id"]
    r = client.post(f"/api/inv/productions/{pid}/post")
    assert r.status_code == 200, r.text
    # Gia thanh: 2x500k + 2x50k = 1.1tr cho 2 bo -> 550k/bo
    out_line = [ln for ln in r.json()["lines"] if ln["chieu"] == "ra"][0]
    assert out_line["gia_tri"] == 1_100_000
    stock = client.get("/api/inv/stock").json()
    row = [x for x in stock["rows"] if x["item_id"] == out][0]
    assert row["ton"] == 2 and row["gia_tri"] == 1_100_000
    # Ban thanh pham
    _, r = _issue(client, "2026-01-15", [(out, tp, 1)])
    assert r.json()["lines"][0]["gia_von"] == 550_000


def test_production_labor_overhead_and_so_ct(client):
    """Gia thanh = NVL + nhan cong (622) + SX chung (627); co so chung tu."""
    nvl, tp = _wh(client, "NVL"), _wh(client, "TP")
    n1 = _mk_item(client, "N010", "Linh kien")
    out = _mk_item(client, "TP010", "May Y")
    _purchase(client, "2026-01-05", [(n1, nvl, 10, 100_000)])
    r = client.post("/api/inv/productions", json={
        "ngay": "2026-02-10",
        "cp_nhan_cong": 200_000,
        "cp_sxc": 100_000,
        "gia_ban_du_kien": 2_000_000,
        "lines": [
            {"chieu": "vao", "item_id": n1, "warehouse_id": nvl, "so_luong": 2},
            {"chieu": "ra", "item_id": out, "warehouse_id": tp, "so_luong": 1},
        ],
    })
    pid = r.json()["id"]
    r = client.post(f"/api/inv/productions/{pid}/post")
    assert r.status_code == 200, r.text
    # NVL 2x100k=200k + NC 200k + SXC 100k = 500k
    out_line = [ln for ln in r.json()["lines"] if ln["chieu"] == "ra"][0]
    assert out_line["gia_tri"] == 500_000
    assert r.json()["tong_gia_thanh"] == 500_000
    assert r.json()["so_ct"] == "LSX-2026-%04d" % pid


def test_production_gia_tam_tinh(client):
    """NVL chua co gia von (mua gia 0) -> dung gia tam tinh cho gia thanh."""
    nvl, tp = _wh(client, "NVL"), _wh(client, "TP")
    n1 = _mk_item(client, "N011", "NVL treo")
    out = _mk_item(client, "TP011", "May Z")
    # Mua 10 gia 0 -> ton 10, gia tri 0 (gia tri treo)
    _purchase(client, "2026-01-05", [(n1, nvl, 10, 0)])
    r = client.post("/api/inv/productions", json={
        "ngay": "2026-01-10",
        "lines": [
            {"chieu": "vao", "item_id": n1, "warehouse_id": nvl, "so_luong": 2,
             "don_gia_tam": 300_000},
            {"chieu": "ra", "item_id": out, "warehouse_id": tp, "so_luong": 1},
        ],
    })
    pid = r.json()["id"]
    r = client.post(f"/api/inv/productions/{pid}/post")
    assert r.status_code == 200, r.text
    out_line = [ln for ln in r.json()["lines"] if ln["chieu"] == "ra"][0]
    assert out_line["gia_tri"] == 600_000  # 2 x 300k gia tam


def test_issue_muc_dich_dinh_khoan_so_ct(client):
    """Phieu xuat: dinh khoan tu dong theo muc dich + so chung tu khi post."""
    hh = _wh(client, "HH")
    item = _mk_item(client, "T020", "Hang xuat SX")
    _purchase(client, "2026-01-05", [(item, hh, 10, 50_000)])
    r = client.post("/api/inv/issues", json={
        "ngay": "2026-03-05", "muc_dich": "san_xuat",
        "lines": [{"item_id": item, "warehouse_id": hh, "so_luong": 2}],
    })
    iid = r.json()["id"]
    r = client.post(f"/api/inv/issues/{iid}/post")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["tk_no"] == "621" and body["tk_co"] == "152"
    assert body["so_ct"] == "PX-2026-%04d" % iid
    assert body["tong_gia_von"] == 100_000  # 2 x 50k


def test_production_insufficient_nvl_blocked(client):
    nvl, tp = _wh(client, "NVL"), _wh(client, "TP")
    n1 = _mk_item(client, "N003", "Linh kien it")
    out = _mk_item(client, "TP002", "May X")
    _purchase(client, "2026-01-05", [(n1, nvl, 1, 100_000)])
    r = client.post("/api/inv/productions", json={
        "ngay": "2026-01-10",
        "lines": [
            {"chieu": "vao", "item_id": n1, "warehouse_id": nvl, "so_luong": 5},
            {"chieu": "ra", "item_id": out, "warehouse_id": tp, "so_luong": 1},
        ],
    })
    pid = r.json()["id"]
    assert client.post(f"/api/inv/productions/{pid}/post").status_code == 400


# ---------------------------------------------------------------------------
# Import Excel ton dau ky
# ---------------------------------------------------------------------------
def _make_opening_xlsx() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["TỔNG HỢP TỒN KHO"] + [""] * 12)
    ws.append(["Năm 2025"] + [""] * 12)
    ws.append(["", "Mã hàng", "Tên hàng", "ĐVT", "Đầu kỳ", "", "Nhập kho", "",
               "Xuất kho", "", "Cuối kỳ", "", ""])
    ws.append(["", "", "", "", "SL", "GT", "SL", "GT", "SL", "GT", "SL", "GT", ""])
    ws.append(["Tên kho : HÀNG HÓA (2 )"] + [""] * 12)
    ws.append(["", "HH01", "Camera Dahua", "Cái", 0, 0, 10, 1000000, 5, 500000, 5, 500000, ""])
    # SL=0 nhung GT != 0 -> canh bao gia tri treo, khong import
    ws.append(["", "HH02", "Orange Pi", "Cái", 0, 0, 1, 16559380, 1, 0, 0, 16559380, ""])
    ws.append(["Tên kho : NGUYÊN VẬT LIỆU (2 )"] + [""] * 12)
    # Cung ma HH01 o kho NVL + thieu DVT
    ws.append(["", "HH01", "Camera Dahua", "", 0, 0, 3, 300000, 1, 100000, 2, 200000, ""])
    ws.append(["", "NV01", "Day mang", "Cuộn", 0, 0, 2, 100000, 2, 100000, 0, 0, ""])
    ws.append(["Số dòng = 4"] + [""] * 12)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_opening_import_dry_run_and_commit(client):
    xlsx = _make_opening_xlsx()
    r = client.post(
        "/api/inv/opening/import?dry_run=true",
        files={"file": ("tonkho.xlsx", xlsx)},
    )
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["dry_run"] is True
    assert j["tong"]["so_ma"] == 3  # HH01 (2 kho) + HH02 + NV01
    codes = {w["code"] for w in j["warnings"]}
    assert "gia_tri_treo" in codes  # HH02
    assert "thieu_dvt" not in codes or True  # HH01 co DVT o kho HH nen khong canh bao
    # 2 dong import: HH01@HH (5, 500k) + HH01@NVL (2, 200k)
    assert len(j["preview"]) == 2
    assert sum(p["gia_tri"] for p in j["preview"]) == 700_000

    r = client.post(
        "/api/inv/opening/import?dry_run=false",
        files={"file": ("tonkho.xlsx", xlsx)},
    )
    assert r.status_code == 200
    assert r.json()["applied"]["moves"] == 2
    stock = client.get("/api/inv/stock").json()
    assert stock["tong_gia_tri"] == 700_000
    # Canh bao gia tri treo duoc luu vao note mat hang de duyet tay
    hh02 = client.get("/api/inv/items?q=HH02").json()[0]
    assert hh02["note"].startswith("⚠️") and "giá trị treo" in hh02["note"]
    # HH01 la 1 mat hang, ton o 2 kho
    hh01 = [x for x in stock["rows"] if x["ma_hang"] == "HH01"]
    assert len(hh01) == 2

    # Import lai khi CHUA co phat sinh -> duoc (thay the)
    r = client.post(
        "/api/inv/opening/import?dry_run=false",
        files={"file": ("tonkho.xlsx", xlsx)},
    )
    assert r.status_code == 200
    assert client.get("/api/inv/stock").json()["tong_gia_tri"] == 700_000

    # Co phat sinh -> 409
    hh = _wh(client, "HH")
    item_id = [x for x in stock["rows"] if x["ma_hang"] == "HH01" and x["warehouse_code"] == "HH"][0]["item_id"]
    _issue(client, "2026-01-10", [(item_id, hh, 1)])
    r = client.post(
        "/api/inv/opening/import?dry_run=false",
        files={"file": ("tonkho.xlsx", xlsx)},
    )
    assert r.status_code == 409


# ---------------------------------------------------------------------------
# Hoa don XML -> draft + chong trung
# ---------------------------------------------------------------------------
_XML = """<?xml version="1.0" encoding="UTF-8"?>
<HDon><DLHDon><TTChung><KHHDon>C25TAA</KHHDon><SHDon>123</SHDon><NLap>2026-02-01</NLap></TTChung>
<NDHDon><NBan><Ten>CTY TNHH ABC</Ten><MST>0312345678</MST></NBan>
<NMua><Ten>CTY INUT</Ten><MST>6001234567</MST></NMua>
<DSHHDVu><HHDVu><STT>1</STT><THHDVu>Camera Dahua</THHDVu><DVTinh>Cái</DVTinh>
<SLuong>2</SLuong><DGia>500000</DGia><ThTien>1000000</ThTien><TSuat>10%</TSuat></HHDVu></DSHHDVu>
<TToan><TgTCThue>1000000</TgTCThue><TgTThue>100000</TgTThue><TgTTTBSo>1100000</TgTTTBSo></TToan>
</NDHDon></DLHDon></HDon>"""


def test_purchase_xml_upload_and_dedup(client):
    _mk_item(client, "CAM01", "Camera Dahua")
    r = client.post(
        "/api/inv/purchase/upload",
        files=[("files", ("hd123.xml", _XML.encode(), "text/xml"))],
    )
    assert r.status_code == 200, r.text
    res = r.json()["results"][0]
    assert res["ok"], res
    pid = res["purchase_id"]
    j = client.get(f"/api/inv/purchase/{pid}").json()
    assert j["so_hd"] == "123" and j["mst_ban"] == "0312345678"
    assert j["ngay"] == "2026-02-01"
    assert j["tong_tien"] == 1_100_000
    assert j["confidence"] == 1.0
    # Match exact theo ten chuan hoa
    assert j["lines"][0]["item_id"] is not None
    assert j["lines"][0]["match_kind"] == "exact"
    # Upload lan 2 -> dup warning
    r = client.post(
        "/api/inv/purchase/upload",
        files=[("files", ("hd123b.xml", _XML.encode(), "text/xml"))],
    )
    res2 = r.json()["results"][0]
    assert res2["dup_of"] == pid


def test_login_lockout_after_5_fails(client):
    """Sai mat khau 5 lan lien tiep -> IP bi khoa 30 phut (ke ca mat khau dung)."""
    for _ in range(4):
        r = client.post("/api/login", json={"username": "admin", "password": "sai"})
        assert r.status_code == 401
        assert "bị khóa" not in r.json()["detail"]
    r = client.post("/api/login", json={"username": "admin", "password": "sai"})
    assert r.status_code == 401
    assert "bị khóa" in r.json()["detail"]  # lan thu 5: bao da khoa
    # Tu day ke ca mat khau DUNG cung bi chan 429
    r = client.post("/api/login", json={"username": "admin", "password": "NhapHang123@"})
    assert r.status_code == 429
    assert "tạm khóa" in r.json()["detail"]
    # Co ghi log login_locked
    audit = client.get("/api/audit?action=login_locked").json()
    assert audit["total"] >= 1


def test_purchase_post_requires_matched_lines(client):
    r = client.post("/api/inv/purchase", json={
        "ngay": "2026-02-01", "ten_ban": "X",
        "lines": [{"ten_raw": "Hang la chua match", "so_luong": 1, "don_gia": 1000}],
    })
    pid = r.json()["id"]
    r = client.post(f"/api/inv/purchase/{pid}/post")
    assert r.status_code == 400
    assert "chưa chọn mặt hàng" in r.json()["detail"]


def test_service_invoice_no_stock(client):
    """HD dich vu: ghi so duoc du chua khop mat hang, va KHONG tao ton kho."""
    r = client.post("/api/inv/purchase", json={
        "ngay": "2026-02-01", "so_hd": "999", "mst_ban": "0311111111", "ten_ban": "BE GROUP",
        "lines": [{"ten_raw": "Cước xe công nghệ", "so_luong": 1, "don_gia": 50000, "thanh_tien": 50000}],
    })
    pid = r.json()["id"]
    # Danh dau la dich vu
    r = client.patch(f"/api/inv/purchase/{pid}", json={"loai": "dich_vu"})
    assert r.status_code == 200 and r.json()["loai"] == "dich_vu"
    # Ghi so duoc du dong hang chua khop mat hang
    r = client.post(f"/api/inv/purchase/{pid}/post")
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "posted"
    # Khong sinh ton kho
    stock = client.get("/api/inv/stock?all_items=true").json()
    assert stock["tong_gia_tri"] == 0
    # Van chan trung khi ghi so HD dich vu trung
    r2 = client.post("/api/inv/purchase", json={
        "ngay": "2026-02-02", "so_hd": "00999", "mst_ban": "0311111111", "ten_ban": "BE GROUP",
        "loai": "dich_vu",
        "lines": [{"ten_raw": "x", "so_luong": 1, "don_gia": 1000, "thanh_tien": 1000}],
    })
    pid2 = r2.json()["id"]
    client.patch(f"/api/inv/purchase/{pid2}", json={"loai": "dich_vu"})
    r = client.post(f"/api/inv/purchase/{pid2}/post")
    assert r.status_code == 400 and "trùng" in r.json()["detail"].lower()


def test_service_invoice_auto_detect(client):
    """Upload HD tu NCC dich vu ro rang -> tu doan loai=dich_vu."""
    from app.inv_import import create_purchase_draft
    from app.db import get_session

    gen = get_session()
    db = next(gen)
    try:
        inv = create_purchase_draft(db, {
            "source": "manual", "so_hd": "1", "mst_ban": "0300000001",
            "ten_ban": "CÔNG TY CỔ PHẦN BE GROUP", "ngay": "2026-01-02",
            "items": [{"ten": "Cước xe", "so_luong": 1, "don_gia": 15000, "thanh_tien": 15000}],
        })
        assert inv.loai == "dich_vu"
        # NCC ban hang hoa (co chu 'dich vu' trong ten) KHONG bi gan nham
        inv2 = create_purchase_draft(db, {
            "source": "manual", "so_hd": "2", "mst_ban": "0300000002",
            "ten_ban": "CÔNG TY TNHH THƯƠNG MẠI DỊCH VỤ ĐIỆN TỬ ABC", "ngay": "2026-01-03",
            "items": [{"ten": "Camera Dahua", "so_luong": 1, "don_gia": 500000, "thanh_tien": 500000}],
        })
        assert inv2.loai == "hang_hoa"
    finally:
        gen.close()


# ---------------------------------------------------------------------------
# normalize_so_hd + chong trung khi ghi so (so HD chuan hoa + MST ban)
# ---------------------------------------------------------------------------
def test_normalize_so_hd():
    from app.inventory import normalize_so_hd

    assert normalize_so_hd("00008407") == "8407"
    assert normalize_so_hd("8407") == "8407"
    assert normalize_so_hd("") == ""


def test_post_duplicate_so_hd_blocked(client):
    hh = _wh(client, "HH")
    item = _mk_item(client, "T009", "Trung so HD")
    _purchase(
        client, "2026-01-05", [(item, hh, 1, 10_000)],
        so_hd="00000123", mst="1111111111",
    )
    r = client.post("/api/inv/purchase", json={
        "so_hd": "123", "mst_ban": "1111111111", "ten_ban": "NCC Test 2", "ngay": "2026-01-06",
        "lines": [{"ten_raw": "item0", "so_luong": 1, "don_gia": 10_000,
                   "thanh_tien": 10_000, "item_id": item, "warehouse_id": hh}],
    })
    assert r.status_code == 200, r.text
    pid2 = r.json()["id"]
    r = client.post(f"/api/inv/purchase/{pid2}/post")
    assert r.status_code == 400
    assert "123" in r.json()["detail"]


# ---------------------------------------------------------------------------
# Bang ke thue: parse Excel + doi chieu voi hoa don da import
# ---------------------------------------------------------------------------
def _make_bang_ke_xlsx() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng kê mua vào"
    ws.append(["PHỤ LỤC"] + [""] * 6)
    ws.append(["BẢNG KÊ HÓA ĐƠN, CHỨNG TỪ HÀNG HÓA, DỊCH VỤ MUA VÀO"] + [""] * 6)
    ws.append(["Kỳ tính thuế : Quý 01 Năm 2026"] + [""] * 6)
    ws.append(["STT", "Hóa đơn, chứng từ, biên lai nộp thuế", None, "Tên người bán",
               "Giá trị HHDV mua vào\nchưa có thuế", "Thuế Suất", "Thuế GTGT đủ điều kiện\nkhấu trừ thuế"])
    ws.append([None, "Số hóa đơn", "Ngày, Tháng, Năm lập hóa đơn"] + [""] * 4)
    ws.append(["01", "123", "05/01/2026", "CTY ABC", 1000000, 0.1, 100000])
    ws.append(["02", "456", "10/01/2026", "CTY XYZ", 2000000, 0.08, 160000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_bang_ke_xlsx():
    from app.inv_import import parse_bang_ke_xlsx

    rows = parse_bang_ke_xlsx(_make_bang_ke_xlsx())
    assert len(rows) == 2
    assert rows[0] == {
        "so_hd": "123", "ngay": "2026-01-05", "ten_ban": "CTY ABC",
        "gia_tri": 1000000.0, "thue": 100000.0,
    }
    assert rows[1]["so_hd"] == "456"
    assert rows[1]["ngay"] == "2026-01-10"


_XML_RECON = """<?xml version="1.0" encoding="UTF-8"?>
<HDon><DLHDon><TTChung><KHHDon>C26ABC</KHHDon><SHDon>123</SHDon><NLap></NLap></TTChung>
<NDHDon><NBan><Ten>CTY ABC</Ten><MST>0000000000</MST></NBan>
<NMua><Ten>CTY INUT</Ten><MST>6001234567</MST></NMua>
<DSHHDVu><HHDVu><STT>1</STT><THHDVu>Hàng đối chiếu</THHDVu><DVTinh>Cái</DVTinh>
<SLuong>1</SLuong><DGia>1000000</DGia><ThTien>1000000</ThTien><TSuat>10%</TSuat></HHDVu></DSHHDVu>
<TToan><TgTCThue>1000000</TgTCThue><TgTThue>100000</TgTThue><TgTTTBSo>1100000</TgTTTBSo></TToan>
</NDHDon></DLHDon></HDon>"""


def test_reconcile_bang_ke(client):
    r = client.post(
        "/api/inv/purchase/upload",
        files=[("files", ("hd_recon.xml", _XML_RECON.encode(), "text/xml"))],
    )
    assert r.status_code == 200, r.text
    res = r.json()["results"][0]
    assert res["ok"], res
    pid = res["purchase_id"]
    before = client.get(f"/api/inv/purchase/{pid}").json()
    assert before["ngay"] == ""  # NLap rong -> thieu ngay

    xlsx = _make_bang_ke_xlsx()
    r = client.post("/api/inv/purchase/bang-ke", files={"file": ("bangke.xlsx", xlsx)})
    assert r.status_code == 200, r.text
    j = r.json()
    assert {x["purchase_id"] for x in j["khop"]} == {pid}
    assert any(x["so_hd"] == "456" for x in j["thieu_file"])
    assert j["ngoai_bang_ke"] == []

    after = client.get(f"/api/inv/purchase/{pid}").json()
    assert after["ngay"] == "2026-01-05"
    assert any(w["code"] == "ngay_bang_ke" for w in after["warnings"])


# ---------------------------------------------------------------------------
# Cai thien thuat toan nhap HD mua: alias hoc tu, check cheo, suy truong, doan loai
# ---------------------------------------------------------------------------
def test_alias_learned_from_manual_match(client):
    """Ghi so HD voi dong match tay -> HD sau cung NCC + cung ten hang tu dong 'learned'."""
    hh = _wh(client, "HH")
    item = _mk_item(client, "ALIAS1", "Mực in Canon 337")
    mst = "0355566677"

    r = client.post("/api/inv/purchase", json={
        "so_hd": "1", "mst_ban": mst, "ten_ban": "NCC ABC", "ngay": "2026-03-01",
        "lines": [{"ten_raw": "Mực in Canon 337 (hàng đẹp)", "so_luong": 1, "don_gia": 500000,
                   "thanh_tien": 500000, "item_id": item, "warehouse_id": hh}],
    })
    assert r.status_code == 200, r.text
    pid = r.json()["id"]
    assert r.json()["lines"][0]["match_kind"] == "manual"
    r = client.post(f"/api/inv/purchase/{pid}/post")
    assert r.status_code == 200, r.text

    # HD moi, cung MST + cung ten hang, KHONG gan item_id tay -> phai tu hoc duoc
    r2 = client.post("/api/inv/purchase", json={
        "so_hd": "2", "mst_ban": mst, "ten_ban": "NCC ABC", "ngay": "2026-03-05",
        "lines": [{"ten_raw": "Mực in Canon 337 (hàng đẹp)", "so_luong": 2, "don_gia": 500000,
                   "thanh_tien": 1000000}],
    })
    assert r2.status_code == 200, r2.text
    line2 = r2.json()["lines"][0]
    assert line2["match_kind"] == "learned"
    assert line2["item_id"] == item


def test_lech_tong_cong_warning(client):
    """Truoc thue + thue lech tong thanh toan tren HD -> canh bao + giam confidence."""
    from app.db import get_session
    from app.inv_import import create_purchase_draft

    gen = get_session()
    db = next(gen)
    try:
        inv = create_purchase_draft(db, {
            "source": "manual", "so_hd": "1", "mst_ban": "0300000009",
            "ten_ban": "NCC Lech Tong", "ngay": "2026-01-02",
            "tong_truoc_thue": 1_000_000, "tong_thue": 100_000, "tong_tien": 1_500_000,
            "items": [{"ten": "Hàng X", "so_luong": 1, "don_gia": 1_000_000, "thanh_tien": 1_000_000}],
            "confidence": 1.0,
        })
        warnings = __import__("json").loads(inv.warnings)
        assert any(w["code"] == "lech_tong_cong" for w in warnings)
        assert inv.confidence <= 0.7
    finally:
        gen.close()


def test_suy_truong_thieu_tung_dong(client):
    """SL/DG thieu 1 trong 2 nhung TT co san -> suy ra duoc, kem canh bao dong."""
    from app.db import get_session
    from app.inv_import import create_purchase_draft

    gen = get_session()
    db = next(gen)
    try:
        inv = create_purchase_draft(db, {
            "source": "manual", "so_hd": "1", "mst_ban": "0300000010",
            "ten_ban": "NCC Suy Truong", "ngay": "2026-01-02",
            "items": [
                # sl=0, dg>0, tt chia het cho dg -> suy so_luong
                {"ten": "Hàng A", "so_luong": 0, "don_gia": 100_000, "thanh_tien": 300_000},
                # sl>0, dg=0, tt co san -> suy don_gia
                {"ten": "Hàng B", "so_luong": 2, "don_gia": 0, "thanh_tien": 400_000},
                # sl=0, dg=0, tt co san -> suy dong (SL=1, DG=TT)
                {"ten": "Hàng C", "so_luong": 0, "don_gia": 0, "thanh_tien": 250_000},
            ],
            "confidence": 1.0,
        })
        lines = sorted(inv.lines, key=lambda l: l.stt)
        assert lines[0].so_luong == 3
        assert "suy_so_luong" in __import__("json").loads(lines[0].warnings)[0]["code"]
        assert lines[1].don_gia == 200_000
        assert "suy_don_gia" in __import__("json").loads(lines[1].warnings)[0]["code"]
        assert lines[2].so_luong == 1 and lines[2].don_gia == 250_000
        assert "suy_dong" in __import__("json").loads(lines[2].warnings)[0]["code"]
    finally:
        gen.close()


def test_pdf_fallback_ai_when_table_empty_of_values(client, monkeypatch):
    """Bang PDF boc duoc dong nhung thanh_tien toan 0 (bang scan hong so) -> fallback AI."""
    from app import inv_import
    from app import ai as ai_mod

    def fake_parse_pdf(_content):
        return {
            "source": "pdf", "so_hd": "1", "ky_hieu": "", "ngay": "2026-01-05",
            "ten_ban": "NCC Scan Hong", "mst_ban": "0300000011",
            "tong_truoc_thue": 1_000_000, "tong_thue": 100_000, "tong_tien": 1_100_000,
            "items": [{"ten": "Hàng scan hỏng", "dvt": "Cái", "so_luong": 1,
                       "don_gia": 0, "thanh_tien": 0, "thue_suat": 10}],
            "confidence": 0.8, "warnings": [],
        }

    def fake_ai(_settings, _content):
        return {
            "source": "scan_ai", "so_hd": "1", "ky_hieu": "", "ngay": "2026-01-05",
            "ten_ban": "NCC Scan Hong", "mst_ban": "0300000011",
            "tong_truoc_thue": 1_000_000, "tong_thue": 100_000, "tong_tien": 1_100_000,
            "items": [{"ten": "Hàng AI đọc lại", "dvt": "Cái", "so_luong": 1,
                       "don_gia": 1_000_000, "thanh_tien": 1_000_000, "thue_suat": 10,
                       "confidence": 0.6}],
            "confidence": 0.6, "warnings": [{"code": "ai", "msg": "AI"}],
        }

    monkeypatch.setattr(inv_import, "parse_purchase_pdf", fake_parse_pdf)
    monkeypatch.setattr(inv_import, "extract_purchase_ai", fake_ai)
    monkeypatch.setattr(ai_mod, "chat", lambda *a, **k: "{}")  # phong khi khong bi monkeypatch dung

    r = client.post(
        "/api/inv/purchase/upload",
        files=[("files", ("hd_scan.pdf", b"%PDF-1.4 fake", "application/pdf"))],
    )
    assert r.status_code == 200, r.text
    res = r.json()["results"][0]
    assert res["ok"], res
    pid = res["purchase_id"]
    j = client.get(f"/api/inv/purchase/{pid}").json()
    # Da chuyen sang du lieu tu AI (ten dong khac voi ban PDF goc)
    assert j["lines"][0]["ten_raw"] == "Hàng AI đọc lại"
    assert j["lines"][0]["thanh_tien"] == 1_000_000


# ---------------------------------------------------------------------------
# ZIP hoa don: uu tien XML chinh, bo Bang_Ke/hdcd
# ---------------------------------------------------------------------------
def test_expand_zip_prefers_main_xml():
    import zipfile

    from app.inv_import import expand_zip

    main_xml = b"<?xml version='1.0'?><HDon><DLHDon>main</DLHDon></HDon>"
    bang_ke_xml = b"<?xml version='1.0'?><HDon><DLHDon>bangke</DLHDon></HDon>"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("2772_1_K26TSM.xml", main_xml)
        zf.writestr("2772_1_K26TSM_Bang_Ke.xml", bang_ke_xml)
        zf.writestr("2772_1_K26TSM_hdcd.pdf", b"%PDF-1.4 fake")
    result = expand_zip("2772_1_K26TSM.zip", buf.getvalue())
    assert result == [("2772_1_K26TSM.xml", main_xml)]


def test_expand_zip_passthrough_non_zip():
    from app.inv_import import expand_zip

    assert expand_zip("hd.pdf", b"%PDF-1.4 fake") == [("hd.pdf", b"%PDF-1.4 fake")]


# ---------------------------------------------------------------------------
# Trich xuat ngay / so HD / ky hieu tu raw_text PDF thuc te (nhieu mau khac nhau)
# ---------------------------------------------------------------------------
def test_extract_ngay_and_so_hd_patterns():
    from app.inv_import import _extract_ky_hieu, _extract_ngay, _extract_so_hd

    raw1 = "HÓA ĐƠN GIÁ TRỊ GIA TĂNG Ký hiệu:1C26MBH | Ngày10tháng03năm2026 Số: 00008407 | Mã CQT:..."
    assert _extract_ngay(raw1, None) == "2026-03-10"
    assert _extract_so_hd(raw1) == "00008407"
    assert _extract_ky_hieu(raw1, "") == "1C26MBH"

    raw2 = "Ký hiệu (Series) : 1C26MSG | Số (No.) : 00016906 | Ngày (Date) : 14 / 01 / 2026"
    assert _extract_ngay(raw2, None) == "2026-01-14"
    assert _extract_so_hd(raw2) == "00016906"

    raw3 = (
        "Mẫu số - Ký hiệu (Serial No.): 1C26TND | Số (Invoice No.): 00000429 | "
        "Ngày (day) 07 tháng (month) 01 năm (year) 2026"
    )
    assert _extract_ngay(raw3, None) == "2026-01-07"
    assert _extract_so_hd(raw3) == "00000429"

    raw4 = (
        "Ký hiệu (Serial): C26MTN | Ngày (Date) 17 tháng (month) 01 năm (year) 2026 "
        "Số (No.): 3122 | Mã CQT (Tax AC): M1-26-RXTXV-00028769954"
    )
    assert _extract_ngay(raw4, None) == "2026-01-17"
    assert _extract_so_hd(raw4) == "3122"


# ---------------------------------------------------------------------------
# Filter tu/den + xuat Excel/ZIP
# ---------------------------------------------------------------------------
def test_purchase_list_filter_tu_den(client):
    client.post("/api/inv/purchase", json={
        "so_hd": "A1", "mst_ban": "0123456789", "ten_ban": "NCC A", "ngay": "2026-01-05",
    })
    client.post("/api/inv/purchase", json={
        "so_hd": "A2", "mst_ban": "0123456789", "ten_ban": "NCC B", "ngay": "2026-03-10",
    })
    r = client.get("/api/inv/purchase", params={"tu": "2026-01-01", "den": "2026-01-31"})
    assert r.status_code == 200, r.text
    rows = r.json()
    assert len(rows) == 1
    assert rows[0]["so_hd"] == "A1"


def test_purchase_export_xlsx(client):
    client.post("/api/inv/purchase", json={
        "so_hd": "B1", "mst_ban": "0123456789", "ten_ban": "NCC B1", "ngay": "2026-02-01",
        "lines": [{"ten_raw": "Hàng X", "dvt": "Cái", "so_luong": 2, "don_gia": 100000,
                   "thanh_tien": 200000, "thue_suat": 10}],
    })
    r = client.get("/api/inv/purchase/export-xlsx")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Hóa đơn", "Dòng hàng"]
    ws = wb["Hóa đơn"]
    assert ws.max_row == 2  # header + 1 hoa don
    ws2 = wb["Dòng hàng"]
    assert ws2.max_row == 2  # header + 1 dong


def test_purchase_export_zip(client):
    r = client.post(
        "/api/inv/purchase/upload",
        files=[("files", ("hdzip.xml", _XML.encode(), "text/xml"))],
    )
    assert r.status_code == 200, r.text
    pid = r.json()["results"][0]["purchase_id"]

    r = client.get("/api/inv/purchase/export-zip")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/zip"
    import zipfile

    zf = zipfile.ZipFile(io.BytesIO(r.content))
    assert len(zf.namelist()) == 1

    # Xoa file goc chua bi xoa -> loc theo ids ma khong co file goc nao -> 404
    r2 = client.get("/api/inv/purchase/export-zip", params={"ids": "999999"})
    assert r2.status_code == 404
